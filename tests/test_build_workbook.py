"""Tests for app/enrichment/workbook.py.

Three structural tests per the build spec:
  1. NOC rows with DIN = "Not Applicable" / blank are excluded from Sheet 1.
  2. Sheet 1 rows are sorted by DIN ascending.
  3. Sheet 2 (Generic Submissions) is standalone — never joined to Sheet 1.
"""
from __future__ import annotations

import io
from datetime import date as _date
from datetime import datetime, timezone
from typing import Any

import pytest

from app.models import DrugRecord, SearchMetadata, SearchResponse, SourceResult


# ── shared fixture builder ────────────────────────────────────────────────────

def _make_response(
    dpd_records: list[DrugRecord] | None = None,
    noc_records: list[DrugRecord] | None = None,
    gsur_records: list[DrugRecord] | None = None,
    query: str = "alpelisib",
) -> SearchResponse:
    # When noc_records is not supplied, auto-generate one NOC stub per DPD DIN so
    # that build_sheet1 (which now sources its DIN universe from NOC) still produces
    # rows.  Tests that specifically exercise NOC logic pass noc_records explicitly.
    if noc_records is None and dpd_records:
        from app.models import DrugRecord as _DR
        noc_records = [
            _DR(
                source="NOC", din=r.din, brand_name=r.brand_name,
                source_specific={
                    "noc_date": "2019-01-01",
                    "submission_type": "NDS",
                    "submission_class": "New",
                    "reason_for_supplement": None,
                    "therapeutic_class": "Test",
                },
            )
            for r in dpd_records
            if r.din and r.din.strip().lower() not in {"", "not applicable", "n/a", "na", "none"}
        ]
    sources = []
    if dpd_records is not None:
        sources.append(SourceResult(source="DPD", status="ok", records=dpd_records))
    if noc_records is not None:
        sources.append(SourceResult(source="NOC", status="ok", records=noc_records))
    if gsur_records is not None:
        sources.append(SourceResult(source="GenericSubmissions", status="ok", records=gsur_records))
    return SearchResponse(
        metadata=SearchMetadata(
            query=query,
            field="ingredient",
            timestamp=datetime.now(timezone.utc).isoformat(),
        ),
        sources=sources,
    )


def _dpd(din: str, brand: str = "BRAND", strength: str = "50 mg") -> DrugRecord:
    return DrugRecord(
        source="DPD",
        din=din,
        brand_name=brand,
        company="Novartis",
        ingredient="alpelisib",
        strength=strength,
        all_ingredients=["alpelisib"],
    )


def _noc(din: str, brand: str = "BRAND", submission_type: str = "NDS") -> DrugRecord:
    return DrugRecord(
        source="NOC",
        din=din,
        brand_name=brand,
        company="Novartis",
        ingredient="alpelisib",
        source_specific={
            "noc_date": "2019-05-24",
            "submission_type": submission_type,
            "submission_class": "New",
            "reason_for_supplement": None,
            "therapeutic_class": "Oncology",
        },
    )


def _gsur(ingredient: str = "alpelisib") -> DrugRecord:
    return DrugRecord(
        source="GenericSubmissions",
        ingredient=ingredient,
        company="GenericCo",
        source_specific={"therapeutic_area": "Oncology", "date_accepted": "2022/06"},
    )


# ── Test 1: NOC "Not Applicable" DINs excluded from Sheet 1 ──────────────────

def test_noc_not_applicable_din_excluded():
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[
            _noc("02498014"),          # valid — should appear
            _noc("Not Applicable"),    # must be excluded
            _noc(""),                  # blank — must be excluded
            _noc("N/A"),               # N/A — must be excluded
        ],
    )
    df = build_sheet1(response)

    assert not df.empty, "Sheet 1 should have rows"
    dins_in_sheet = set(df["din"].astype(str).str.strip())
    # "Not Applicable", "", "N/A" must not appear
    assert "Not Applicable" not in dins_in_sheet
    assert "" not in dins_in_sheet
    assert "N/A" not in dins_in_sheet
    # Valid DIN from both NOC and DPD should be present
    assert "02498014" in dins_in_sheet


def test_noc_not_applicable_only_gives_empty_sheet1():
    """If all NOC rows have N/A DINs and there are no DPD records, Sheet 1 is empty."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        noc_records=[_noc("Not Applicable"), _noc("N/A")],
    )
    df = build_sheet1(response)
    assert df.empty or len(df) == 0


# ── Test 2: Sheet 1 rows sorted by DIN ascending ─────────────────────────────

def test_sheet1_din_sorted_ascending():
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[
            _dpd("02498022"),
            _dpd("00012345"),
            _dpd("02498014"),
        ],
    )
    df = build_sheet1(response)
    dins = list(df["din"].astype(str))
    assert dins == sorted(dins), f"DINs should be sorted ascending, got: {dins}"


def test_sheet1_single_row_still_sorted():
    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet1(response)
    assert len(df) == 1
    assert df.iloc[0]["din"] == "02498014"


# ── Test 3: Sheet 2 is standalone — no DIN column, only GSUR data ────────────

def test_sheet2_is_standalone_no_din_column():
    from app.enrichment.workbook import build_sheet2

    response = _make_response(
        dpd_records=[_dpd("02498014")],
        gsur_records=[_gsur("alpelisib"), _gsur("alpelisib hydrochloride")],
    )
    df = build_sheet2(response)
    assert "din" not in df.columns, "Sheet 2 must not have a DIN column"
    assert "medicinal_ingredient" in df.columns


def test_sheet2_filtered_to_queried_ingredient():
    from app.enrichment.workbook import build_sheet2

    response = _make_response(
        gsur_records=[
            _gsur("alpelisib"),
            _gsur("alpelisib hydrochloride"),
            _gsur("metformin"),          # unrelated — must NOT appear
            _gsur("metformin hydrochloride"),  # unrelated — must NOT appear
        ],
        query="alpelisib",
    )
    df = build_sheet2(response)
    ings = list(df["medicinal_ingredient"].str.lower())
    assert all("alpelisib" in i for i in ings), (
        f"Sheet 2 should only contain alpelisib submissions, got: {ings}"
    )
    assert not any("metformin" in i for i in ings), "metformin rows must be filtered out"


def test_sheet2_empty_when_no_gsur_source():
    from app.enrichment.workbook import build_sheet2

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet2(response)
    assert len(df) == 0


# ── Test 4: build_workbook produces valid XLSX bytes ─────────────────────────

def test_build_workbook_returns_xlsx():
    from app.enrichment.workbook import build_workbook

    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[_noc("02498014")],
        gsur_records=[_gsur()],
    )
    xlsx = build_workbook(response)

    # Should be a valid ZIP/XLSX (starts with PK magic bytes)
    assert xlsx[:2] == b"PK", "XLSX should start with PK (ZIP magic bytes)"

    # Verify both sheets exist
    import zipfile
    with zipfile.ZipFile(io.BytesIO(xlsx)) as zf:
        names = zf.namelist()
    # openpyxl encodes sheet names in xl/worksheets/
    assert any("sheet" in n.lower() for n in names), "No worksheet files found in XLSX"


def test_build_workbook_sheet1_has_patent_columns(tmp_path):
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    store_mod.upsert_patent("02498014", "2709025", "2008-12-10", "2014-08-26", "2028-12-10")

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet1(response)
    assert "patent_count" in df.columns, "patent_count missing"
    assert "patent_number" in df.columns, "patent_number missing"
    assert "patent_grant_date" in df.columns, "patent_grant_date missing"
    assert "patent_expiry_date" in df.columns, "patent_expiry_date missing"
    # Old wide columns must be absent
    assert "patent_1_number" not in df.columns, "stale patent_1_number column present"
    assert "patent_numbers" not in df.columns, "stale patent_numbers column present"
    assert "all_patents_detail" not in df.columns, "stale all_patents_detail column present"


def test_build_workbook_sheet1_has_labeling_columns(tmp_path):
    import time
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    # Add minimal labeling data so labeling columns are not dropped by the
    # all-empty cleanup (Change 2).  These sentinels are real values ("Not in PM"
    # / "No PM available"), so the columns must be kept.
    store_mod.upsert_labeling("02498014", {
        "color": "pink", "shape": "round", "ph": "Not stated",
        "needs_ocr": 0, "has_unverified": 0, "drug_code": 99001,
        "fetched_at": time.time(),
    })

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet1(response)
    for col in ("color", "shape", "ph"):
        assert col in df.columns, f"Expected labeling column '{col}' in Sheet 1"


# ── Test 5: patent block aggregation ─────────────────────────────────────────

def test_patent_aggregation(tmp_path):
    """_aggregate_patents_latest selects the latest-expiry patent."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    store_mod.upsert_patent("02498014", "2709025", "2008-12-10", "2014-08-26", "2028-12-10")
    store_mod.upsert_patent("02498014", "2900000", "2015-01-01", "2020-03-01", "2035-01-01")

    from app.enrichment.workbook import _aggregate_patents_latest
    agg = _aggregate_patents_latest("02498014")

    assert agg["patent_count"] == 2
    # 2900000 has the later expiry (2035 vs 2028) — must be selected
    assert agg["patent_number"] == "2900000", (
        f"Expected 2900000 (latest expiry), got {agg['patent_number']!r}"
    )
    assert agg["patent_grant_date"] == "2020-03-01"
    assert agg["patent_expiry_date"] == "2035-01-01"


# ── Test 6: fail-loud guard — error-vs-no_results distinction ─────────────────

def _make_response_with_noc_error(query: str = "metformin") -> "SearchResponse":
    """Response where NOC has status=error, DPD has results."""
    from app.models import SourceResult
    sources = [
        SourceResult(source="DPD", status="ok", records=[_dpd("02242974")]),
        SourceResult(source="NOC", status="error", error_message="test forced error"),
        SourceResult(source="GenericSubmissions", status="no_results"),
        SourceResult(source="PatentRegister", status="no_results"),
    ]
    return SearchResponse(
        metadata=SearchMetadata(
            query=query,
            field="ingredient",
            timestamp="2026-01-01T00:00:00+00:00",
        ),
        sources=sources,
    )


def test_build_workbook_with_source_errors_adds_warning_sheet():
    """allow_partial=True path: source_errors causes a '⚠ Source Status' sheet to appear."""
    import openpyxl
    from app.enrichment.workbook import build_workbook

    response = _make_response_with_noc_error()
    xlsx = build_workbook(response, source_errors={"NOC": "test forced error"})

    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    sheet_names = wb.sheetnames
    assert any("Status" in name for name in sheet_names), (
        f"Expected a Source Status warning sheet; got sheets: {sheet_names}"
    )


def test_build_workbook_no_source_errors_no_warning_sheet():
    """Default path (source_errors=None): no warning sheet is added."""
    import openpyxl
    from app.enrichment.workbook import build_workbook

    response = _make_response(dpd_records=[_dpd("02498014")], noc_records=[_noc("02498014")])
    xlsx = build_workbook(response, source_errors=None)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert len(wb.sheetnames) == 2, (
        f"Expected exactly 2 sheets with no errors; got: {wb.sheetnames}"
    )


def test_build_workbook_no_results_source_does_not_block():
    """A no_results source (Patent, GSUR) must not add a warning sheet."""
    import openpyxl
    from app.enrichment.workbook import build_workbook

    response = _make_response(dpd_records=[_dpd("02498014")])
    # no_results is NOT an error — source_errors dict should be empty → no sheet
    xlsx = build_workbook(response, source_errors=None)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert len(wb.sheetnames) == 2


async def test_export_refuses_with_409_on_source_error(mock_noc, mock_dpd, mock_gsur, mock_patent_register):
    """HTTP 409 when a source is in error and allow_partial is not set."""
    from fastapi.testclient import TestClient
    from app.main import app
    from unittest.mock import AsyncMock, patch
    from app.models import SourceResult

    forced_error = SourceResult(
        source="NOC", status="error", error_message="forced test error"
    )

    with patch("app.main.search_noc", new=AsyncMock(return_value=forced_error)), \
         patch("app.main.enrich_patents", new=AsyncMock(return_value={})), \
         patch("app.main.enrich_labeling_batch", new=AsyncMock(return_value={})):
        client = TestClient(app)
        resp = client.get("/api/export?q=metformin&field=ingredient")
        assert resp.status_code == 409, f"Expected 409, got {resp.status_code}: {resp.text}"
        body = resp.json()
        assert "NOC" in body.get("detail", "")
        assert "allow_partial" in body.get("detail", "")


async def test_export_allow_partial_builds_with_warning(mock_noc, mock_dpd, mock_gsur, mock_patent_register):
    """allow_partial=true: builds a workbook with the source-status warning sheet."""
    from fastapi.testclient import TestClient
    import openpyxl
    from app.main import app
    from unittest.mock import AsyncMock, patch
    from app.models import SourceResult

    forced_error = SourceResult(
        source="NOC", status="error", error_message="forced test error"
    )

    with patch("app.main.search_noc", new=AsyncMock(return_value=forced_error)), \
         patch("app.main.enrich_patents", new=AsyncMock(return_value={})), \
         patch("app.main.enrich_labeling_batch", new=AsyncMock(return_value={})):
        client = TestClient(app)
        resp = client.get("/api/export?q=metformin&field=ingredient&allow_partial=true")
        assert resp.status_code == 200, f"Expected 200 with allow_partial; got {resp.status_code}"
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert any("Status" in name for name in wb.sheetnames), (
            f"Expected Source Status sheet; got: {wb.sheetnames}"
        )


# ── Test 7: SNDS/SANDS filtering ─────────────────────────────────────────────

def test_snds_rows_excluded_from_sheet1():
    """NOC records with SNDS submission types are filtered; the DIN only appears if
    it also has a non-SNDS NOC entry or appears with a different DIN in NOC."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[
            _noc("02498014", submission_type="NDS"),                             # keep
            _noc("02498022", submission_type="Supplement to a New Drug Submission (SNDS)"),  # filtered
        ],
    )
    df = build_sheet1(response)

    # 02498014 (NDS) must appear; 02498022 (SNDS-only) is excluded from NOC universe → not in sheet
    dins = set(df["din"].astype(str))
    assert "02498014" in dins
    assert "02498022" not in dins, (
        "DIN with only an SNDS NOC entry is filtered from NOC universe and must not appear"
    )
    row_nds = df[df["din"] == "02498014"].iloc[0]
    assert row_nds["noc_submission_type"] == "NDS"


def test_sands_rows_excluded_from_sheet1():
    """SANDS-only DIN has no NOC entry after filtering → excluded from sheet entirely."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498022")],
        noc_records=[
            _noc("02498022", submission_type="Supplement to an Abbreviated New Drug Submission (SANDS)"),
        ],
    )
    df = build_sheet1(response)
    # SANDS is filtered → NOC has no valid DIN → sheet is empty
    assert df.empty or "02498022" not in set(df["din"].astype(str)), (
        "DIN with only a SANDS NOC entry must be excluded from Sheet 1"
    )


def test_dpd_only_din_excluded_from_sheet():
    """A DPD DIN with no matching NOC record is excluded from Sheet 1.

    Under the NOC-authoritative rule, DPD-only DINs do not appear in Sheet 1;
    they appear in the exclusion list (build_exclusion_list) instead.
    """
    from app.enrichment.workbook import build_sheet1, build_exclusion_list

    # DPD record for 02498022 but no NOC record for it
    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[_noc("02498014")],  # only 02498014 has NOC
    )
    df = build_sheet1(response)

    # 02498014 appears with real NOC data; 02498022 is excluded (DPD-only)
    dins = set(df["din"].astype(str))
    assert "02498014" in dins
    assert "02498022" not in dins, "DPD-only DIN must not appear in Sheet 1"

    row_with_noc = df[df["din"] == "02498014"].iloc[0]
    assert row_with_noc["noc_submission_type"] == "NDS"
    assert row_with_noc["noc_date"] == "2019-05-24"

    # 02498022 must appear in the exclusion list
    excl = build_exclusion_list(response)
    assert "02498022" in set(excl["din"].values)


def test_noc_only_din_excluded_from_sheet():
    """A NOC DIN with no matching DPD record is excluded from Sheet 1.

    NOC-only DINs (present in NOC but absent from DPD, e.g. 02272113 / 02272121)
    carry no DPD product data and must not appear in the export.
    """
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014")],
        noc_records=[
            _noc("02498014"),    # present in both NOC and DPD → appears
            _noc("02272113"),    # NOC-only → must be excluded
            _noc("02272121"),    # NOC-only → must be excluded
        ],
    )
    df = build_sheet1(response)

    dins = set(df["din"].astype(str))
    assert dins == {"02498014"}, (
        f"Only DINs present in BOTH NOC and DPD should appear, got: {dins}"
    )
    assert "02272113" not in dins
    assert "02272121" not in dins


def test_noc_only_column_values_after_filtering():
    """After SNDS/SANDS filtering, only NDS and ANDS appear in noc_submission_type.

    SNDS-only DINs are excluded from the NOC universe and therefore absent from
    Sheet 1 entirely — 'No NOC record' never appears since all included DINs come
    from NOC.
    """
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022"), _dpd("02498030")],
        noc_records=[
            _noc("02498014", submission_type="NDS"),
            _noc("02498022", submission_type="ANDS"),
            _noc("02498030", submission_type="Supplement to a New Drug Submission (SNDS)"),
        ],
    )
    df = build_sheet1(response)

    # 02498030 (SNDS-only) is excluded; only 02498014 and 02498022 appear
    assert set(df["din"].astype(str)) == {"02498014", "02498022"}
    allowed = {"NDS", "ANDS"}
    for val in df["noc_submission_type"].dropna():
        assert str(val) in allowed, (
            f"Unexpected noc_submission_type in Sheet 1: {val!r}"
        )


# ── Test 8: No PM available sentinel ─────────────────────────────────────────

def test_no_pm_available_constant_exported():
    """NO_PM_AVAILABLE constant must be importable and distinct from NOT_IN_PM."""
    from app.enrichment.labeling import NO_PM_AVAILABLE, NOT_IN_PM
    assert NO_PM_AVAILABLE != NOT_IN_PM
    assert NO_PM_AVAILABLE == "No PM available"


# ── Test 9: Change 2 — all-empty column cleanup ──────────────────────────────

def test_empty_patents_columns_present_with_none(tmp_path):
    """When no DIN has patents, patent_number/grant_date/expiry_date are None.

    patent_count stays (0 is a real integer, not empty); patent_number etc. are
    in _NEVER_DROP_COLS so they must also be kept.
    """
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014"), _dpd("02498022")])
    df = build_sheet1(response)

    assert "patent_count" in df.columns, "patent_count must stay even when all counts are 0"
    assert "patent_number" in df.columns, "patent_number must stay (in _NEVER_DROP_COLS)"
    # Old wide-format columns must not exist
    assert "patent_1_number" not in df.columns
    assert "patent_1_filing_date" not in df.columns


def test_patent_number_present_for_din_with_patent(tmp_path):
    """A DIN with a patent gets the latest-expiry patent_number; a DIN without gets None."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    store_mod.upsert_patent("02498014", "2709025", "2008-12-10", "2014-08-26", "2028-12-10")

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014"), _dpd("02498022")])
    df = build_sheet1(response)

    assert "patent_number" in df.columns
    row_with = df[df["din"] == "02498014"].iloc[0]
    row_without = df[df["din"] == "02498022"].iloc[0]
    assert row_with["patent_number"] == "2709025"
    assert row_without["patent_number"] is None or str(row_without["patent_number"]) in (
        "None", "nan", ""
    )


def test_noc_columns_always_present(tmp_path):
    """NOC columns are never pruned — they are in _NEVER_DROP_COLS."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014"), _dpd("02498022")])
    df = build_sheet1(response)

    for col in ("noc_date", "noc_submission_type", "reason_for_supplement", "submission_class"):
        assert col in df.columns, f"{col} must always be present (in _NEVER_DROP_COLS)"


def test_single_nonempty_row_prevents_column_drop(tmp_path):
    """A column with even one non-empty value must not be dropped,
    even if all other rows are None."""
    import time
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    # Give one DIN a labeling color; other DIN has nothing
    store_mod.upsert_labeling("02498014", {
        "color": "blue", "needs_ocr": 0, "has_unverified": 0,
        "drug_code": 99001, "fetched_at": time.time(),
    })
    # 02498022 has no labeling → color=None for that row

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014"), _dpd("02498022")])
    df = build_sheet1(response)

    assert "color" in df.columns, (
        "color must be kept: 02498014 has a non-empty value even though 02498022 is None"
    )


def test_export_never_produces_old_sheet_names():
    """The /api/export endpoint must never produce the old multi-sheet format."""
    import openpyxl
    from app.enrichment.workbook import build_workbook

    response = _make_response(
        dpd_records=[_dpd("02498014")],
        noc_records=[_noc("02498014")],
        gsur_records=[_gsur()],
    )
    xlsx = build_workbook(response)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    old_names = {
        "Drug Product Database (DPD)",
        "Generic Submissions Under Review",
        "Notice of Compliance (NOC)",
        "Patent Register",
        "Combined",
        "By Combination",
        "Search Metadata",
    }
    for sheet_name in wb.sheetnames:
        assert sheet_name not in old_names, (
            f"Old sheet name '{sheet_name}' must not appear in the new two-sheet workbook"
        )
    assert set(wb.sheetnames) == {"DPD + NOC + Patents", "Generic Submissions"}, (
        f"Expected exactly the two new sheet names; got: {wb.sheetnames}"
    )


# ── Fix 1: Active-only patents ───────────────────────────────────────────────
#
# Verification anchors (as-of 2026-06-01):
#   ENDOMETRIN  02334992  expiry 2026-11-24  → active  → count 1, details shown
#   BIJUVA      02505223  expiry 2032-11-21  → active  → count 1, details shown
#   INPROSUB    02515504  expiry 2025-09-14  → expired → count "all patents expired", cells blank

_AS_OF = _date(2026, 6, 1)


def test_active_patent_shows_details(tmp_path):
    """A patent whose expiry is after as_of is counted and its details are populated."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    # ENDOMETRIN-like: single patent expiring 2026-11-24 (active as of 2026-06-01)
    store_mod.upsert_patent("02334992", "2334992P", "2006-11-24", "2011-03-15", "2026-11-24")

    from app.enrichment.workbook import _aggregate_patents_latest
    agg = _aggregate_patents_latest("02334992", as_of=_AS_OF)

    assert agg["patent_count"] == 1, f"Expected 1 active patent, got {agg['patent_count']!r}"
    assert agg["patent_number"] == "2334992P"
    assert agg["patent_expiry_date"] == "2026-11-24"
    assert agg["patent_grant_date"] == "2011-03-15"


def test_bijuva_active_patent(tmp_path):
    """BIJUVA-like patent (expiry 2032) is active as of 2026-06-01."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    store_mod.upsert_patent("02505223", "2505223P", "2012-11-21", "2018-05-01", "2032-11-21")

    from app.enrichment.workbook import _aggregate_patents_latest
    agg = _aggregate_patents_latest("02505223", as_of=_AS_OF)

    assert agg["patent_count"] == 1
    assert agg["patent_number"] == "2505223P"
    assert agg["patent_expiry_date"] == "2032-11-21"


def test_expired_patent_sentinel(tmp_path):
    """INPROSUB-like patent (expiry 2025-09-14) is expired as of 2026-06-01.

    patent_count must be the exact string 'all patents expired';
    detail cells must be None.
    """
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    # INPROSUB: single patent expired before as_of
    store_mod.upsert_patent("02515504", "2515504P", "2005-09-14", "2010-03-01", "2025-09-14")

    from app.enrichment.workbook import _aggregate_patents_latest
    agg = _aggregate_patents_latest("02515504", as_of=_AS_OF)

    assert agg["patent_count"] == "all patents expired", (
        f"Expected sentinel 'all patents expired', got {agg['patent_count']!r}"
    )
    assert agg["patent_number"] is None, "patent_number must be None when all expired"
    assert agg["patent_grant_date"] is None, "patent_grant_date must be None when all expired"
    assert agg["patent_expiry_date"] is None, "patent_expiry_date must be None when all expired"


def test_mixed_patents_only_active_counted(tmp_path):
    """Two patents: one expired, one active → count = 1, details from the active one."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    store_mod.upsert_patent("02334992", "EXPIRED1", "2000-01-01", "2005-01-01", "2020-01-01")
    store_mod.upsert_patent("02334992", "ACTIVE1",  "2010-01-01", "2016-01-01", "2030-01-01")

    from app.enrichment.workbook import _aggregate_patents_latest
    agg = _aggregate_patents_latest("02334992", as_of=_AS_OF)

    assert agg["patent_count"] == 1
    assert agg["patent_number"] == "ACTIVE1"
    assert agg["patent_expiry_date"] == "2030-01-01"


def test_no_patents_stored_gives_zero(tmp_path):
    """DIN with no patents stored → patent_count = 0 (not the expired sentinel)."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    from app.enrichment.workbook import _aggregate_patents_latest
    agg = _aggregate_patents_latest("00000000", as_of=_AS_OF)

    assert agg["patent_count"] == 0, (
        "No-patent DIN must have count 0, not the 'all patents expired' sentinel"
    )
    assert agg["patent_number"] is None


def test_expired_patents_in_sheet1(tmp_path):
    """Sheet 1 shows the 'all patents expired' sentinel for an expired DIN."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    store_mod.upsert_patent("02515504", "2515504P", "2005-09-14", "2010-03-01", "2025-09-14")

    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[DrugRecord(
            source="DPD", din="02515504", brand_name="INPROSUB",
            company="AnyPharma", ingredient="progesterone",
            strength="25 MG", dosage_form="Solution",
        )],
    )
    df = build_sheet1(response, as_of=_AS_OF)
    row = df[df["din"] == "02515504"].iloc[0]

    assert row["patent_count"] == "all patents expired"
    assert row["patent_number"] is None or str(row["patent_number"]) in ("None", "nan", "")
    assert row["patent_expiry_date"] is None or str(row["patent_expiry_date"]) in ("None", "nan", "")


# ── Fix 2: Solution strength normalization ───────────────────────────────────

def test_solution_strength_normalized():
    """'25 MG' with dosage_form 'Solution' → '25 MG/ML' (INPROSUB anchor)."""
    from app.enrichment.workbook import _normalize_solution_strength
    assert _normalize_solution_strength("25 MG", "Solution") == "25 MG/ML"


def test_solution_strength_already_has_denominator():
    """Strengths that already carry a denominator are left unchanged."""
    from app.enrichment.workbook import _normalize_solution_strength
    assert _normalize_solution_strength("25 MG/ML", "Solution") == "25 MG/ML"
    assert _normalize_solution_strength("10 MG/5 ML", "Solution") == "10 MG/5 ML"


def test_non_solution_strength_unchanged():
    """Non-Solution forms are not touched regardless of unit."""
    from app.enrichment.workbook import _normalize_solution_strength
    assert _normalize_solution_strength("25 MG", "Capsule") == "25 MG"
    assert _normalize_solution_strength("25 MG", "Suspension") == "25 MG"
    assert _normalize_solution_strength("25 MG", "Gel") == "25 MG"


def test_inprosub_solution_strength_in_sheet1(tmp_path):
    """INPROSUB row: form='Solution', raw strength='25 MG' → sheet shows '25 MG/ML'."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[DrugRecord(
            source="DPD", din="02515504", brand_name="INPROSUB",
            company="AnyPharma", ingredient="progesterone",
            strength="25 MG", dosage_form="Solution",
        )],
    )
    df = build_sheet1(response)
    row = df[df["din"] == "02515504"].iloc[0]
    assert row["strength"] == "25 MG/ML", (
        f"INPROSUB Solution strength should be '25 MG/ML', got {row['strength']!r}"
    )


def test_capsule_strength_unchanged_in_sheet1(tmp_path):
    """A Capsule row with '50 MG' must not have /ML appended."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[DrugRecord(
            source="DPD", din="02498014", brand_name="PIQRAY",
            company="Novartis", ingredient="alpelisib",
            strength="50 MG", dosage_form="Tablet",
        )],
    )
    df = build_sheet1(response)
    row = df[df["din"] == "02498014"].iloc[0]
    assert row["strength"] == "50 MG", (
        f"Tablet strength must not be modified, got {row['strength']!r}"
    )


# ── Fix 3: DIN-first column order + DIN / SKU Name headers ──────────────────

def test_din_is_first_column(tmp_path):
    """DIN must be the first column in Sheet 1."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet1(response)
    assert df.columns[0] == "din", (
        f"First column must be 'din', got {df.columns[0]!r}"
    )


def test_xlsx_din_header_is_all_caps(tmp_path):
    """The 'din' column must appear as 'DIN' (all caps) in the XLSX header row."""
    import openpyxl
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    from app.enrichment.workbook import build_workbook

    response = _make_response(dpd_records=[_dpd("02498014")], noc_records=[_noc("02498014")])
    xlsx = build_workbook(response)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb["DPD + NOC + Patents"]
    headers = [cell.value for cell in ws[1]]
    assert "DIN" in headers, f"Expected 'DIN' header in XLSX; got headers: {headers}"
    assert "Din" not in headers, "Old title-cased 'Din' must not appear"


def test_xlsx_ingredient_header_is_sku_name(tmp_path):
    """The 'ingredient' column must appear as 'SKU Name' in the XLSX header row."""
    import openpyxl
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    from app.enrichment.workbook import build_workbook

    response = _make_response(dpd_records=[_dpd("02498014")], noc_records=[_noc("02498014")])
    xlsx = build_workbook(response)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb["DPD + NOC + Patents"]
    headers = [cell.value for cell in ws[1]]
    assert "SKU Name" in headers, f"Expected 'SKU Name' header in XLSX; got: {headers}"
    assert "Ingredient" not in headers, "Old 'Ingredient' header must not appear"


def test_xlsx_din_is_first_column_in_file(tmp_path):
    """Column A in the XLSX must be DIN."""
    import openpyxl
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    from app.enrichment.workbook import build_workbook

    response = _make_response(dpd_records=[_dpd("02498014")], noc_records=[_noc("02498014")])
    xlsx = build_workbook(response)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb["DPD + NOC + Patents"]
    col_a_header = ws.cell(row=1, column=1).value
    assert col_a_header == "DIN", (
        f"Column A header must be 'DIN', got {col_a_header!r}"
    )


# ── Fix 4: NOC column alignment (DEPO-PROVERA anchor) ───────────────────────

def _noc_multi(din: str, date: str, reason: Optional[str], sub_type: str = "NDS") -> DrugRecord:
    return DrugRecord(
        source="NOC", din=din,
        brand_name="BRAND", company="Pfizer",
        ingredient="medroxyprogesterone",
        source_specific={
            "noc_date": date,
            "submission_type": sub_type,
            "submission_class": "New",
            "reason_for_supplement": reason,
            "therapeutic_class": "Hormones",
        },
    )


def test_depo_provera_reason_for_supplement_aligned():
    """DEPO-PROVERA anchor: 3 NOC dates, first reason blank, next two 'MANUFACTURER NAME CHANGE'.

    After the fix, reason_for_supplement must have 3 lines (not 2),
    aligned with noc_date, submission_class, etc.
    """
    from app.enrichment.workbook import build_sheet1

    noc_records = [
        _noc_multi("00030848", "1997-05-29", None),                          # original → blank reason
        _noc_multi("00030848", "2001-02-13", "MANUFACTURER NAME CHANGE"),
        _noc_multi("00030848", "2003-09-15", "MANUFACTURER NAME CHANGE"),
    ]
    response = _make_response(
        dpd_records=[DrugRecord(
            source="DPD", din="00030848", brand_name="DEPO-PROVERA",
            company="Pfizer", ingredient="medroxyprogesterone acetate",
            strength="150 MG/ML", dosage_form="Suspension",
        )],
        noc_records=noc_records,
    )
    df = build_sheet1(response)
    row = df[df["din"] == "00030848"].iloc[0]

    noc_date_lines = str(row["noc_date"]).split("\n")
    reason_lines = str(row["reason_for_supplement"]).split("\n")
    sub_class_lines = str(row["submission_class"]).split("\n")

    # All three NOC columns must have the same line count.
    assert len(noc_date_lines) == 3, (
        f"noc_date should have 3 lines, got {len(noc_date_lines)}: {noc_date_lines!r}"
    )
    assert len(reason_lines) == len(noc_date_lines), (
        f"reason_for_supplement line count ({len(reason_lines)}) must match "
        f"noc_date line count ({len(noc_date_lines)})"
    )
    assert len(sub_class_lines) == len(noc_date_lines), (
        f"submission_class line count ({len(sub_class_lines)}) must match "
        f"noc_date line count ({len(noc_date_lines)})"
    )

    # Line 1 reason must be blank (original approval has no supplement reason).
    assert reason_lines[0] == "", (
        f"Line 1 of reason_for_supplement must be blank (original NOC), got {reason_lines[0]!r}"
    )
    # Lines 2 and 3 must carry the reason.
    assert reason_lines[1] == "MANUFACTURER NAME CHANGE"
    assert reason_lines[2] == "MANUFACTURER NAME CHANGE"

    # Dates must align correctly.
    assert "1997-05-29" in noc_date_lines[0]
    assert "2001-02-13" in noc_date_lines[1]
    assert "2003-09-15" in noc_date_lines[2]


def test_noc_column_line_counts_match_multi_noc():
    """For any multi-NOC DIN, all five per-NOC columns must have equal line counts."""
    from app.enrichment.workbook import build_sheet1

    noc_cols = [
        "noc_date", "reason_for_supplement", "submission_class",
        "noc_submission_type", "noc_therapeutic_class",
    ]

    # Three NOC records: mixed blank/non-blank across different fields
    noc_records = [
        _noc_multi("02498014", "2019-01-01", None,                     "NDS"),
        _noc_multi("02498014", "2021-06-01", "LABEL UPDATE",           "NDS"),
        _noc_multi("02498014", "2023-03-15", "MANUFACTURER NAME CHANGE", "NDS"),
    ]
    response = _make_response(
        dpd_records=[_dpd("02498014")],
        noc_records=noc_records,
    )
    df = build_sheet1(response)
    row = df[df["din"] == "02498014"].iloc[0]

    line_counts = {
        col: len(str(row[col]).split("\n")) if row[col] is not None else 0
        for col in noc_cols
        if col in df.columns and row[col] is not None
    }

    # All columns with data must have the same line count.
    unique_counts = set(line_counts.values())
    assert len(unique_counts) == 1, (
        f"All per-NOC columns must have equal line counts; got {line_counts}"
    )
