"""Regression test for the IQVIA quarter-over-quarter comparison.

Anchored to the committed fixture pair in tests/fixtures/iqvia/diff/ (regenerate
with tests/scripts/build_iqvia_diff_fixture.py).  The pair is fully synthetic and
deliberately mismatched across the two files — CSV vs xlsx, YYYY/MM vs MM/YYYY
date order, a decoy Pivot sheet — so this exercises the whole canonical path:
generalized parse → collapse → latest-MAT resolution → normalised-identity diff.

Every expected number below is hand-computed from the fixture rows; see the
builder script's docstring for the per-product rationale.
"""
from pathlib import Path

import pytest
import pandas as pd

from app.enrichment.iqvia_diff import compare_iqvia, build_diff_workbook

_DIFF_DIR = Path(__file__).parent / "fixtures" / "iqvia" / "diff"
_OLD = _DIFF_DIR / "old_extract.csv"
_NEW = _DIFF_DIR / "new_extract.xlsx"


@pytest.fixture(scope="module")
def diff():
    if not (_OLD.exists() and _NEW.exists()):
        pytest.skip(
            "IQVIA diff fixtures missing — run tests/scripts/build_iqvia_diff_fixture.py"
        )
    return compare_iqvia(_OLD.read_bytes(), _NEW.read_bytes())


def _by_product(df: pd.DataFrame, name: str) -> pd.Series:
    rows = df[df["Product"] == name]
    assert len(rows) == 1, f"expected exactly one {name!r} row, got {len(rows)}"
    return rows.iloc[0]


# ── Latest-MAT resolution (per file, independent) ─────────────────────────────

def test_latest_periods_resolved_per_file(diff):
    # Old CSV latest = 2024/06 (YYYY/MM); new xlsx latest = 12/2024 (MM/YYYY).
    assert diff.old_period == (2024, 6)
    assert diff.new_period == (2024, 12)
    # New is strictly later → no swap/equal-period warnings.
    assert diff.warnings == []


# ── Signal partitioning: exact membership, no leakage ─────────────────────────

def test_entrants_exits_moves_membership(diff):
    assert set(diff.entrants["Product"]) == {"OZEMPIC"}
    assert set(diff.exits["Product"]) == {"ZANTAC", "PEPCID"}
    assert set(diff.moves["Product"]) == {"LIPITOR"}


def test_below_threshold_row_absent_everywhere(diff):
    # GLUCOPHAGE moved +$40k (<$100k) and +500u (<1,000): below both floors.
    for df in (diff.entrants, diff.exits, diff.moves):
        assert "GLUCOPHAGE" not in set(df["Product"]), "below-threshold row leaked into output"


# ── ENTRANT ───────────────────────────────────────────────────────────────────

def test_entrant_values(diff):
    r = _by_product(diff.entrants, "OZEMPIC")
    assert r["Combined Molecule"] == "SEMAGLUTIDE"
    assert int(r["Dollars"]) == 5_000_000
    assert int(r["Units"]) == 200_000
    assert int(r["Ext Units"]) == 200_000


# ── EXITS ──────────────────────────────────────────────────────────────────────

def test_exit_absent_from_new(diff):
    r = _by_product(diff.exits, "ZANTAC")           # gone from the new file entirely
    assert int(r["Dollars"]) == 800_000
    assert int(r["Units"]) == 60_000


def test_exit_zero_latest_in_new(diff):
    # PEPCID's row exists in new but its latest-MAT cells are "-" (0) → still an exit.
    r = _by_product(diff.exits, "PEPCID")
    assert int(r["Dollars"]) == 300_000           # the OLD value is reported


def test_exits_sorted_by_dollars_desc(diff):
    assert list(diff.exits["Product"]) == ["ZANTAC", "PEPCID"]


# ── MATERIAL MOVE (+ identity normalisation) ──────────────────────────────────

def test_material_move_values(diff):
    r = _by_product(diff.moves, "LIPITOR")
    # Old split across two channel rows (1.2M + 0.8M) must have collapsed to 2.0M.
    assert int(r["Dollars Old"]) == 2_000_000
    assert int(r["Dollars New"]) == 2_500_000
    assert int(r["Dollars Δ"]) == 500_000
    assert float(r["Dollars Δ%"]) == 25.0
    assert int(r["Units Old"]) == 100_000
    assert int(r["Units New"]) == 130_000
    assert int(r["Units Δ"]) == 30_000
    assert float(r["Units Δ%"]) == 30.0


def test_material_move_survives_formatting_jitter(diff):
    # Old "PFIZER CANADA ULC"/"LIPITOR 20MG TABLETS" vs new "PFIZER"/"LIPITOR" must
    # fold to ONE identity — never a phantom exit + entrant.
    assert "LIPITOR" not in set(diff.entrants["Product"])
    assert "LIPITOR" not in set(diff.exits["Product"])


# ── Workbook shape ─────────────────────────────────────────────────────────────

def test_workbook_has_four_sheets(diff):
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(build_diff_workbook(diff)), read_only=True)
    assert wb.sheetnames == ["Summary", "New Entrants", "Exits", "Material Moves"]


def test_summary_has_no_extract_date_fields(diff):
    # The per-file MAT periods are reporting periods, not extract/pull dates, so
    # they were removed from the Summary sheet. Assert they cannot silently return:
    # no Summary "Metric" cell may reference an extract date / MAT period.
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(build_diff_workbook(diff)), read_only=True)
    metrics = [
        str(row[0].value or "")
        for row in wb["Summary"].iter_rows(min_row=2, max_col=1)
    ]
    assert not any("extract" in m.lower() or "mat period" in m.lower() for m in metrics), metrics


def test_no_invented_values_pct_blank_when_no_base():
    # A product present in old with zero Dollars but positive Units (so it is a
    # move, not an entrant) must leave Dollars Δ% blank rather than fabricate a %.
    # Build a minimal in-memory pair to assert the contract directly.
    from tests.scripts.build_iqvia_diff_fixture import _ID, _OLD_METRICS, _NEW_METRICS
    import csv as _csv, io as _io

    old_rows = [["Drugstore", "DRUGX", "10MG", "30 TAB", "BRANDX", "TAB ORAL",
                 "ACME", "ACME", "TAB", "ONTARIO",
                 "0", "5,000", "5,000", "0", "5,000", "5,000"]]
    new_rows = [["Drugstore", "DRUGX", "10MG", "30 TAB", "BRANDX", "TAB ORAL",
                 "ACME", "ACME", "TAB", "ONTARIO",
                 0, 5000, 5000, 500000, 9000, 9000]]
    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(_ID + _OLD_METRICS)
    w.writerows(old_rows)
    old_bytes = buf.getvalue().encode("utf-8")
    new_df = pd.DataFrame(new_rows, columns=_ID + _NEW_METRICS)
    nbuf = _io.BytesIO()
    with pd.ExcelWriter(nbuf, engine="openpyxl") as xw:
        new_df.to_excel(xw, sheet_name="data", index=False)
    d = compare_iqvia(old_bytes, nbuf.getvalue())
    r = d.moves[d.moves["Product"] == "BRANDX"].iloc[0]
    assert int(r["Dollars Old"]) == 0
    assert int(r["Dollars New"]) == 500_000
    assert pd.isna(r["Dollars Δ%"])              # 0 base → blank, never faked
    assert float(r["Units Δ%"]) == 80.0          # 5,000 → 9,000
