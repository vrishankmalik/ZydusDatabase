"""Permanent regression suite for the IQVIA quarter-over-quarter compare, anchored
to the TWO REAL IQVIA Canada extracts the feature was built for:

    OLD = IQVIA_Canada_Oct 2025(combined data).csv   (CSV, dates YYYY/MM, latest MAT 2025/06)
    NEW = IQVIA.xlsx                                  (xlsx, Pivot+data sheets,
                                                       dates MM/YYYY, latest MAT 2025/12)

The pair naturally exercises every cross-file hazard the code defends against: the
xlsx Pivot-sheet decoy, the YYYY/MM-vs-MM/YYYY date-order gap, and CSV-vs-xlsx value
formatting. This is corroborated by the shipped output workbook the feature already
produced from this exact pair (iqvia_changes_2025-06_to_2025-12.xlsx: 322 New
Entrants, 278 Exits, 3,320 Material Moves).

Locating the files
------------------
Paths are taken from env vars IQVIA_REAL_OLD / IQVIA_REAL_NEW if set, else the
known on-disk locations, else a search of the user's Desktop/Downloads for the
canonical filenames. If neither file is found the whole module SKIPS (it depends on
data outside the repo) — it never silently passes.

The core guarantee — row-by-row, every output row justified
-----------------------------------------------------------
``_oracle`` re-derives the entire classification INDEPENDENTLY of the production
diff path (compare_iqvia / _aggregate_latest are never called inside the oracle): it
collapses each file, folds to one latest-MAT triple per normalised identity, then
partitions every key in (old ∪ new) into exactly one of entrant / exit / move /
unchanged using the documented rules (present = any non-zero latest metric;
materiality = abs-AND-pct floor on Dollars or Units). The tests then assert the
production output equals the oracle EXACTLY — same entrant/exit/move key sets, every
move's Δ for dollars and units, a complete & disjoint partition of (old ∪ new), and
Summary counts that reconcile to the three signal tabs.

Because the oracle reuses the same normalisers as the code, a handful of anchors are
ALSO hand-verified straight from the raw province/channel rows (summed with plain
pandas, bypassing collapse_iqvia) — these catch any bug the oracle and the code
would share.

NOTE (this sandbox only): importing pandas under Python 3.14 here hangs inside
platform._wmi_query(). Runs use an out-of-repo sitecustomize.py on PYTHONPATH that
forces that WMI call to raise (pandas then falls back to the registry path). The
repo is untouched; on a normal interpreter the suite imports cleanly.
"""
from __future__ import annotations

import io
import os
from pathlib import Path

import pandas as pd
import pytest

from app.config import IQVIA_DIFF_DOLLARS_ABS, IQVIA_DIFF_UNITS_ABS, IQVIA_DIFF_PCT
from app.enrichment.iqvia import (
    parse_iqvia,
    collapse_iqvia,
    detect_metric_columns,
    latest_mat_metrics,
    _pick_data_sheet,
    _parse_mat_period,
    _norm_brand,
    _norm_company,
    _norm_strength,
)
from app.enrichment.iqvia_diff import compare_iqvia, build_diff_workbook


# ── Locate the two real extracts ──────────────────────────────────────────────

_OLD_NAME = "IQVIA_Canada_Oct 2025(combined data).csv"
_NEW_NAME = "IQVIA.xlsx"


def _find(name: str, env: str) -> Path | None:
    env_val = os.environ.get(env)
    if env_val and Path(env_val).is_file():
        return Path(env_val)
    home = Path.home()
    candidates = [
        home / "Downloads" / name,
        home / "Desktop" / name,
        home / "OneDrive - Viona Pharmaceuticals USA INC" / "Desktop" / name,
    ]
    for c in candidates:
        if c.is_file():
            return c
    for base in (home / "Downloads", home / "Desktop",
                 home / "OneDrive - Viona Pharmaceuticals USA INC" / "Desktop"):
        if base.is_dir():
            hit = next((p for p in base.glob(name)), None)
            if hit:
                return hit
    return None


_OLD_PATH = _find(_OLD_NAME, "IQVIA_REAL_OLD")
_NEW_PATH = _find(_NEW_NAME, "IQVIA_REAL_NEW")

pytestmark = pytest.mark.skipif(
    not (_OLD_PATH and _NEW_PATH),
    reason=(
        "Real IQVIA extracts not found. Set IQVIA_REAL_OLD / IQVIA_REAL_NEW to the "
        f"two extracts (old={_OLD_NAME!r}, new={_NEW_NAME!r}) to run this suite."
    ),
)


# ── Expected facts about the real pair (observed, not invented) ───────────────

OLD_PERIOD = (2025, 6)
NEW_PERIOD = (2025, 12)
N_ENTRANTS = 322
N_EXITS = 278
N_MOVES = 3320
N_UNCHANGED = 5948
N_UNION = N_ENTRANTS + N_EXITS + N_MOVES + N_UNCHANGED  # 9868

_ID = ["Combined Molecule", "Product", "Manufacturer", "Strength"]
_METRICS = ["dollars", "units", "ext_units"]


def _identity(mol, prod, mfr, strg) -> tuple:
    """Independent re-implementation of the cross-file identity key.

    Mirrors iqvia_diff._identity (molecule upper+whitespace-collapsed, matcher
    normalisers for brand/company, sorted strength-token set). Reusing the
    normalisers is intentional and explicitly allowed — the hand-verified raw-row
    anchors below guard the bug the oracle and code would otherwise share.
    """
    m = " ".join(str(mol or "").split()).upper()
    return (m, _norm_brand(prod), _norm_company(mfr), tuple(sorted(_norm_strength(strg))))


# ── Module-scoped heavy work: run the production compare + build the oracle once ─

@pytest.fixture(scope="module")
def real_bytes():
    return _OLD_PATH.read_bytes(), _NEW_PATH.read_bytes()


@pytest.fixture(scope="module")
def diff(real_bytes):
    old_b, new_b = real_bytes
    return compare_iqvia(old_b, new_b)


@pytest.fixture(scope="module")
def raw_frames(real_bytes):
    """parse_iqvia output for both files (used by hand-verify + derived edge cases)."""
    old_b, new_b = real_bytes
    return parse_iqvia(old_b), parse_iqvia(new_b)


def _aggregate_oracle(raw: pd.DataFrame) -> tuple:
    """INDEPENDENT aggregation: collapse → one latest-MAT triple per identity.

    Deliberately does NOT call iqvia_diff._aggregate_latest. Returns
    (period, {identity: {dollars, units, ext_units, + raw display fields}}).
    """
    collapsed = collapse_iqvia(raw)
    cols = detect_metric_columns(collapsed)
    period, latest = latest_mat_metrics(cols)
    agg: dict[tuple, dict] = {}
    for _, r in collapsed.iterrows():
        k = _identity(r.get("Combined Molecule"), r.get("Product"),
                      r.get("Manufacturer"), r.get("Strength"))
        rec = agg.get(k)
        if rec is None:
            rec = {"dollars": 0, "units": 0, "ext_units": 0,
                   "Combined Molecule": str(r.get("Combined Molecule") or "").strip(),
                   "Product": str(r.get("Product") or "").strip(),
                   "Manufacturer": str(r.get("Manufacturer") or "").strip(),
                   "Strength": str(r.get("Strength") or "").strip()}
            agg[k] = rec
        for mk in _METRICS:
            c = latest.get(mk)
            if c is not None:
                rec[mk] += int(r.get(c, 0) or 0)
    return period, agg


def _present(rec: dict | None) -> bool:
    return rec is not None and (rec["dollars"] > 0 or rec["units"] > 0 or rec["ext_units"] > 0)


def _material(o: dict, n: dict) -> bool:
    for mk, floor in (("dollars", IQVIA_DIFF_DOLLARS_ABS), ("units", IQVIA_DIFF_UNITS_ABS)):
        d = n[mk] - o[mk]
        if abs(d) < floor:
            continue
        pct = abs(d) / o[mk] if o[mk] else float("inf")
        if pct >= IQVIA_DIFF_PCT:
            return True
    return False


@pytest.fixture(scope="module")
def oracle(raw_frames):
    """Independent partition of (old ∪ new) into entrant/exit/move/unchanged."""
    old_raw, new_raw = raw_frames
    op, oagg = _aggregate_oracle(old_raw)
    np_, nagg = _aggregate_oracle(new_raw)
    ent, ex, mv, unchanged = {}, {}, {}, {}
    for k in set(oagg) | set(nagg):
        o, n = oagg.get(k), nagg.get(k)
        op_, np2 = _present(o), _present(n)
        if np2 and not op_:
            ent[k] = n
        elif op_ and not np2:
            ex[k] = o
        elif op_ and np2 and _material(o, n):
            mv[k] = (o, n)
        else:
            unchanged[k] = (o, n)
    return {"old_period": op, "new_period": np_, "old_agg": oagg, "new_agg": nagg,
            "entrants": ent, "exits": ex, "moves": mv, "unchanged": unchanged}


# ── Helpers to read identity-keyed sets out of the production DataFrames ───────

def _df_keyset(df: pd.DataFrame) -> set:
    return {_identity(r["Combined Molecule"], r["Product"], r["Manufacturer"], r["Strength"])
            for _, r in df.iterrows()}


def _move_row(diff, product: str, strength: str) -> pd.Series:
    m = diff.moves[(diff.moves["Product"] == product) & (diff.moves["Strength"] == strength)]
    assert len(m) == 1, f"expected one move row for {product!r}/{strength!r}, got {len(m)}"
    return m.iloc[0]


def _raw_sum(raw: pd.DataFrame, mol: str, prod: str, mfr: str, strength: str | None = None):
    """Hand-sum the latest-MAT metrics over a product's raw province/channel rows.

    Bypasses collapse_iqvia entirely (plain boolean mask + .sum()), so it catches a
    bug shared between the oracle and the production collapse.
    """
    cols = detect_metric_columns(raw)
    _, latest = latest_mat_metrics(cols)
    m = ((raw["Combined Molecule"].astype(str).str.strip() == mol)
         & (raw["Product"].astype(str).str.strip() == prod)
         & (raw["Manufacturer"].astype(str).str.strip() == mfr))
    if strength is not None:
        m = m & (raw["Strength"].astype(str).str.strip() == strength)
    sub = raw[m]
    return {
        "n_rows": int(len(sub)),
        "dollars": int(sub[latest["dollars"]].sum()),
        "units": int(sub[latest["units"]].sum()),
        "ext_units": int(sub[latest["ext_units"]].sum()),
    }


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║ Core: production output == independent oracle, exactly                        ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class TestRowByRowAgainstOracle:
    def test_periods_resolved_per_file(self, diff, oracle):
        assert diff.old_period == OLD_PERIOD == oracle["old_period"]
        assert diff.new_period == NEW_PERIOD == oracle["new_period"]
        assert diff.reordered is False
        assert diff.warnings == []

    def test_counts_match_oracle_and_shipped_output(self, diff, oracle):
        assert (len(diff.entrants), len(diff.exits), len(diff.moves)) == (N_ENTRANTS, N_EXITS, N_MOVES)
        assert len(oracle["entrants"]) == N_ENTRANTS
        assert len(oracle["exits"]) == N_EXITS
        assert len(oracle["moves"]) == N_MOVES

    def test_entrant_set_exact(self, diff, oracle):
        prod, orc = _df_keyset(diff.entrants), set(oracle["entrants"])
        assert prod - orc == set(), f"entrants only in production: {prod - orc}"
        assert orc - prod == set(), f"entrants only in oracle: {orc - prod}"

    def test_exit_set_exact(self, diff, oracle):
        prod, orc = _df_keyset(diff.exits), set(oracle["exits"])
        assert prod - orc == set(), f"exits only in production: {prod - orc}"
        assert orc - prod == set(), f"exits only in oracle: {orc - prod}"

    def test_move_set_exact(self, diff, oracle):
        prod, orc = _df_keyset(diff.moves), set(oracle["moves"])
        assert prod - orc == set(), f"moves only in production: {prod - orc}"
        assert orc - prod == set(), f"moves only in oracle: {orc - prod}"

    def test_every_move_delta_matches_oracle(self, diff, oracle):
        """Every emitted move row's Δ (dollars & units) equals the oracle's Δ."""
        by_key = {_identity(r["Combined Molecule"], r["Product"], r["Manufacturer"], r["Strength"]): r
                  for _, r in diff.moves.iterrows()}
        assert set(by_key) == set(oracle["moves"])
        for k, (o, n) in oracle["moves"].items():
            r = by_key[k]
            assert int(r["Dollars Δ"]) == n["dollars"] - o["dollars"]
            assert int(r["Units Δ"]) == n["units"] - o["units"]
            assert int(r["Dollars Old"]) == o["dollars"] and int(r["Dollars New"]) == n["dollars"]
            assert int(r["Units Old"]) == o["units"] and int(r["Units New"]) == n["units"]

    def test_every_emitted_row_is_justified(self, diff, oracle):
        """For each output row, assert WHY it belongs (the classification rule)."""
        oa, na = oracle["old_agg"], oracle["new_agg"]
        for k in _df_keyset(diff.entrants):
            assert _present(na.get(k)) and not _present(oa.get(k)), f"entrant rule violated: {k}"
        for k in _df_keyset(diff.exits):
            assert _present(oa.get(k)) and not _present(na.get(k)), f"exit rule violated: {k}"
        for k in _df_keyset(diff.moves):
            assert _present(oa.get(k)) and _present(na.get(k)), f"move not present both sides: {k}"
            assert _material(oa[k], na[k]), f"move below materiality gate: {k}"

    def test_partition_complete_and_disjoint(self, oracle):
        """Every key in (old ∪ new) lands in exactly one bucket; nothing lost/double-counted."""
        ent, ex, mv, un = (set(oracle["entrants"]), set(oracle["exits"]),
                           set(oracle["moves"]), set(oracle["unchanged"]))
        union = set(oracle["old_agg"]) | set(oracle["new_agg"])
        # disjoint
        assert ent & ex == set() and ent & mv == set() and ent & un == set()
        assert ex & mv == set() and ex & un == set() and mv & un == set()
        # complete
        assert ent | ex | mv | un == union
        assert len(ent) + len(ex) + len(mv) + len(un) == len(union) == N_UNION

    def test_no_signal_leakage_between_tabs(self, diff):
        """Material Moves contains no entrants/exits; the three tabs are disjoint."""
        e, x, m = _df_keyset(diff.entrants), _df_keyset(diff.exits), _df_keyset(diff.moves)
        assert e & x == set()
        assert e & m == set()
        assert x & m == set()


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║ Hand-verified anchors — summed from raw rows, independent of collapse         ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class TestHandVerifiedAnchors:
    """~6 anchors fully re-summed from the raw province/channel rows by hand."""

    def test_move_eylea(self, diff, oracle, raw_frames):
        old_raw, new_raw = raw_frames
        # Single-strength product → product-level raw sum == per-identity agg.
        o = _raw_sum(old_raw, "AFLIBERCEPT", "EYLEA", "BAYER HEALTHCARE")
        n = _raw_sum(new_raw, "AFLIBERCEPT", "EYLEA", "BAYER HEALTHCARE")
        assert (o["n_rows"], o["dollars"], o["units"]) == (29, 889_036_961, 607_087)
        assert (n["n_rows"], n["dollars"], n["units"]) == (29, 691_232_084, 472_398)
        # classified as a move with exactly this Δ
        k = _identity("AFLIBERCEPT", "EYLEA", "BAYER HEALTHCARE", "2MG/0.05ML")
        assert k in oracle["moves"]
        r = _move_row(diff, "EYLEA", "2MG/0.05ML")
        assert int(r["Dollars Δ"]) == 691_232_084 - 889_036_961 == -197_804_877
        assert int(r["Units Δ"]) == 472_398 - 607_087 == -134_689
        assert float(r["Dollars Δ%"]) == -22.2  # round(-197804877/889036961*100, 1)

    def test_exit_kadcyla(self, diff, oracle, raw_frames):
        old_raw, new_raw = raw_frames
        o = _raw_sum(old_raw, "TRASTUZUMAB EMASTINE", "KADCYLA", "ROCHE")
        n = _raw_sum(new_raw, "TRASTUZUMAB EMASTINE", "KADCYLA", "ROCHE")
        # Present in old (23 rows, two strengths summing to 71,067,293), gone from new.
        assert (o["n_rows"], o["dollars"]) == (23, 71_067_293)
        assert n["n_rows"] == 0 and n["dollars"] == 0
        # the 100MG strength is an exit with its hand-summed old value
        k = _identity("TRASTUZUMAB EMASTINE", "KADCYLA", "ROCHE", "100MG")
        assert k in oracle["exits"]
        assert oracle["exits"][k]["dollars"] == 36_638_448

    def test_entrant_beyfortus(self, diff, oracle, raw_frames):
        old_raw, new_raw = raw_frames
        o = _raw_sum(old_raw, "NIRSEVIMAB", "BEYFORTUS", "SANOFI-PASTEUR")
        n = _raw_sum(new_raw, "NIRSEVIMAB", "BEYFORTUS", "SANOFI-PASTEUR")
        # Absent from old entirely; new product-level total is the sum of its two
        # per-strength entrant rows (66,025,108 + 78,526,542).
        assert o["n_rows"] == 0 and o["dollars"] == 0
        assert (n["n_rows"], n["dollars"]) == (39, 144_551_650)
        k100 = _identity("NIRSEVIMAB", "BEYFORTUS", "SANOFI-PASTEUR", "100MG/ML")
        k50 = _identity("NIRSEVIMAB", "BEYFORTUS", "SANOFI-PASTEUR", "50MG/0.5ML")
        assert k100 in oracle["entrants"] and k50 in oracle["entrants"]
        assert oracle["entrants"][k100]["dollars"] == 66_025_108
        assert oracle["entrants"][k50]["dollars"] == 78_526_542
        assert (oracle["entrants"][k100]["dollars"]
                + oracle["entrants"][k50]["dollars"]) == n["dollars"]

    def test_exit_via_zero_apo_fingolimod(self, oracle, raw_frames):
        """Row EXISTS in new but its latest-MAT cells are 0 → still an EXIT.

        Proves 'present' means non-zero latest sales, not mere row existence.
        """
        old_raw, new_raw = raw_frames
        o = _raw_sum(old_raw, "FINGOLIMOD", "APO-FINGOLIMOD", "APOTEX INC")
        n = _raw_sum(new_raw, "FINGOLIMOD", "APO-FINGOLIMOD", "APOTEX INC")
        assert (o["n_rows"], o["dollars"]) == (10, 112_560)
        assert n["n_rows"] == 8 and (n["dollars"], n["units"], n["ext_units"]) == (0, 0, 0)
        k = _identity("FINGOLIMOD", "APO-FINGOLIMOD", "APOTEX INC", "0.5MG")
        assert k in oracle["exits"]
        # key exists in BOTH aggregates (the new row is present-but-zero)
        assert k in oracle["old_agg"] and k in oracle["new_agg"]
        assert oracle["exits"][k]["dollars"] == 112_560

    def test_collapse_sums_many_province_channel_rows(self, oracle, raw_frames):
        """TARO-CLOBETASOL 0.05% in NEW is summed from 153 raw rows (2 channels,
        9 provinces). The hand-sum must equal the collapsed/aggregated value."""
        _, new_raw = raw_frames
        n = _raw_sum(new_raw, "CLOBETASOL", "TARO-CLOBETASOL", "TARO PHARMACEUTICALS", "0.05%")
        assert n["n_rows"] == 153
        assert (n["dollars"], n["units"]) == (15_714_610, 878_514)
        k = _identity("CLOBETASOL", "TARO-CLOBETASOL", "TARO PHARMACEUTICALS", "0.05%")
        assert oracle["new_agg"][k]["dollars"] == n["dollars"]
        assert oracle["new_agg"][k]["units"] == n["units"]


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║ A. Prompt-1 lock-ins: Summary shape, tab separation, old/new orientation      ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class TestWorkbookAndSummary:
    @pytest.fixture(scope="class")
    def wb(self, diff):
        from openpyxl import load_workbook
        return load_workbook(io.BytesIO(build_diff_workbook(diff)), read_only=True)

    def test_four_sheets(self, wb):
        assert wb.sheetnames == ["Summary", "New Entrants", "Exits", "Material Moves"]

    def test_summary_counts_reconcile_to_tabs(self, wb, diff):
        rows = {str(r[0].value): r[1].value for r in wb["Summary"].iter_rows(min_row=2, max_col=2)}
        assert int(rows["New entrants"]) == len(diff.entrants) == N_ENTRANTS
        assert int(rows["Exits"]) == len(diff.exits) == N_EXITS
        assert int(rows["Material moves"]) == len(diff.moves) == N_MOVES

    def test_summary_keeps_materiality_gate(self, wb):
        metrics = [str(r[0].value or "") for r in wb["Summary"].iter_rows(min_row=2, max_col=1)]
        assert any("Materiality gate — Dollars" in m for m in metrics)
        assert any("Materiality gate — Units" in m for m in metrics)

    def test_summary_has_no_grain_fields(self, wb):
        """Prompt-1 removed per-row grain / extract-date fields from Summary."""
        metrics = [str(r[0].value or "").lower() for r in wb["Summary"].iter_rows(min_row=2, max_col=1)]
        assert not any("extract" in m or "mat period" in m or "grain" in m for m in metrics), metrics


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║ A. Old/new orientation — the silent-flip regression                           ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class TestOldNewOrientation:
    @pytest.fixture(scope="module")
    def diff_reversed(self, real_bytes):
        # Feed the REAL extracts in the WRONG slot order (new in slot 1, old in slot 2).
        old_b, new_b = real_bytes
        return compare_iqvia(new_b, old_b)

    def test_reversed_slots_autocorrect_old_to_new(self, diff_reversed, diff):
        d = diff_reversed
        assert d.reordered is True
        # ordering is by detected latest period, not slot order → identical to forward
        assert d.old_period == OLD_PERIOD
        assert d.new_period == NEW_PERIOD
        assert (len(d.entrants), len(d.exits), len(d.moves)) == (N_ENTRANTS, N_EXITS, N_MOVES)
        assert _df_keyset(d.entrants) == _df_keyset(diff.entrants)
        assert _df_keyset(d.exits) == _df_keyset(diff.exits)

    def test_reversed_surfaces_prominent_banner(self, diff_reversed):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(build_diff_workbook(diff_reversed)), read_only=True)
        first_cells = [str(r[0].value or "") for r in wb["Summary"].iter_rows(min_row=1, max_col=1)]
        assert any("REORDERED" in c.upper() for c in first_cells), first_cells

    def test_same_latest_period_respects_slot_order(self, raw_frames):
        """Two same-latest-period files (derived from real rows) → no swap, slot 1 = old,
        with an informational note (never an error)."""
        old_raw, new_raw = raw_frames
        # Build two extracts that BOTH have latest period 2025/12 by slicing the real
        # NEW file (which is dated 12/YYYY) into two disjoint molecule subsets.
        molsA = ["AFLIBERCEPT", "NIRSEVIMAB"]
        molsB = ["CLOBETASOL", "CLARITHROMYCIN"]
        a = _subset_xlsx(new_raw, molsA)
        b = _subset_xlsx(new_raw, molsB)
        d = compare_iqvia(a, b)
        assert d.reordered is False
        assert d.old_period == d.new_period == NEW_PERIOD
        assert any("same latest period" in w.lower() for w in d.warnings), d.warnings


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║ B. Identity & detection — the real 4-field key                                ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class TestIdentityAndPresence:
    def test_present_means_nonzero_latest_metric(self, oracle):
        """At least one real key sits in both files but is zero on one side, and is
        classified by sales-presence, not row-existence."""
        oa, na = oracle["old_agg"], oracle["new_agg"]
        shared = set(oa) & set(na)
        exit_via_zero = [k for k in shared if _present(oa[k]) and not _present(na[k])]
        entrant_via_zero = [k for k in shared if _present(na[k]) and not _present(oa[k])]
        assert exit_via_zero, "expected at least one present-but-zero exit"
        assert entrant_via_zero, "expected at least one present-but-zero entrant"
        for k in exit_via_zero:
            assert k in oracle["exits"]
        for k in entrant_via_zero:
            assert k in oracle["entrants"]

    def test_entrant_via_zero_named(self, oracle):
        # LOSARTAN / SIVEM PHARMA ULC 25MG: zero latest in old, $163,926 in new.
        k = _identity("LOSARTAN", "LOSARTAN", "SIVEM PHARMA ULC", "25MG")
        assert k in oracle["new_agg"] and k in oracle["old_agg"]
        assert not _present(oracle["old_agg"][k])
        assert oracle["new_agg"][k]["dollars"] == 163_926
        assert k in oracle["entrants"]

    def test_classification_only_new_only_old_both(self, oracle):
        oa, na = oracle["old_agg"], oracle["new_agg"]
        # only in new (no old key at all) → entrant
        k_new_only = _identity("NIRSEVIMAB", "BEYFORTUS", "SANOFI-PASTEUR", "100MG/ML")
        assert k_new_only in na and k_new_only not in oa and k_new_only in oracle["entrants"]
        # only in old → exit
        k_old_only = _identity("TRASTUZUMAB EMASTINE", "KADCYLA", "ROCHE", "100MG")
        assert k_old_only in oa and k_old_only not in na and k_old_only in oracle["exits"]
        # in both, both present → move or unchanged (never entrant/exit)
        k_both = _identity("AFLIBERCEPT", "EYLEA", "BAYER HEALTHCARE", "2MG/0.05ML")
        assert k_both in oa and k_both in na
        assert k_both not in oracle["entrants"] and k_both not in oracle["exits"]


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║ C. Phantom-pair guards — real products that must NOT split into exit+entrant  ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class TestPhantomPairGuards:
    """Each pair is a REAL product whose strength is spelled differently in the two
    files; the normalisers must fold them to one identity (present in both, hence a
    move/unchanged), never a phantom exit + entrant. Confirmed against the diff."""

    # (molecule, product, manufacturer, strength-in-new) for products present in both.
    _PAIRS = [
        ("NAPROXEN", "APO-NAPROXEN", "APOTEX", "500MG"),                       # GM→MG: 0.5GM vs 500MG
        ("TRIPTORELIN", "DECAPEPTYL", "FERRING", "0.1MG"),                     # MCG→MG + /ML strip
        ("VALGANCICLOVIR", "VALCYTE", "CHEPLAPHARM", "50MG"),                  # /ML strip
        ("DIGOXIN", "PMS-DIGOXIN", "PHARMASCIENCE", "0.05MG"),                 # MCG→MG + /ML strip
        ("HYDROCHLOROTHIAZIDE:TELMISARTAN", "MICARDIS PLUS", "BOEHRINGER ING", "80MG/12.5MG"),  # combo order + bare-num
    ]

    @pytest.mark.parametrize("mol,prod,mfr,strg", _PAIRS)
    def test_real_strength_jitter_does_not_split(self, diff, mol, prod, mfr, strg):
        k = _identity(mol, prod, mfr, strg)
        assert k not in _df_keyset(diff.entrants), f"{prod} wrongly split into an entrant"
        assert k not in _df_keyset(diff.exits), f"{prod} wrongly split into an exit"

    # Strength-normaliser rules anchored on the real spellings above.
    def test_strength_gm_to_mg(self):
        assert _norm_strength("0.5GM") == _norm_strength("500MG") == frozenset({"500MG"})

    def test_strength_mcg_and_conc_denominator(self):
        assert _norm_strength("100MCG/ML") == _norm_strength("0.1MG/ML") == frozenset({"0.1MG"})
        assert _norm_strength("50MG/ML") == _norm_strength("50MG") == frozenset({"50MG"})

    def test_strength_combo_order_and_bare_number(self):
        assert _norm_strength("80/12.5MG") == _norm_strength("80MG/12.5MG") == frozenset({"80MG", "12.5MG"})
        assert _norm_strength("12.5MG/80MG") == _norm_strength("80MG/12.5MG")  # order-independent

    def test_strength_decimal_trailing_zero(self):
        assert _norm_strength("10MG") == _norm_strength("10.0MG") == frozenset({"10MG"})

    def test_company_legal_suffix_and_whitespace(self):
        # Real company strings from the files normalise to their bare stem; trivial
        # legal-suffix / whitespace / casing variants collapse together.
        assert _norm_company("APOTEX INC") == _norm_company("Apotex Inc.") == "apotex"
        assert _norm_company("TEVA CANADA LTD") == "teva"
        assert _norm_company("SANDOZ CANADA INC") == _norm_company("  sandoz   canada   inc ") == "sandoz"
        assert _norm_company("VIIV HEALTHCARE ULC") == "viiv"

    def test_company_french_accent_and_suffix(self):
        # Accented French legal forms flatten to ASCII and strip (Limitée→'', Ltée→'').
        assert _norm_company("PRO DOC LIMITÉE") == _norm_company("PRO DOC LIMITEE") == "pro doc"
        assert _norm_company("LABORATOIRE RIVA LTÉE") == _norm_company("LABORATOIRE RIVA LTEE")

    def test_molecule_whitespace_and_casing(self):
        assert _identity("  naproxen  ", "APO-NAPROXEN", "APOTEX", "500MG") \
            == _identity("NAPROXEN", "APO-NAPROXEN", "APOTEX", "500MG")

    def test_brand_trailing_strength_form_stripped(self):
        assert _norm_brand("PROVERA 5MG TABLETS") == "provera"
        assert _norm_brand("APO-NAPROXEN TABLETS") == "apo-naproxen"


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║ C. Known-limitation pin — a corporate rename DOES split (visible, by design)  ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class TestKnownLimitationCorporateRename:
    """BRUKINSA's manufacturer was renamed BeiGene Switzerland GmbH → BeOne Medicines
    GmbH between the two pulls. The names normalise to different words, so the same
    physical product currently surfaces as an EXIT (old company) + ENTRANT (new
    company). This pins the CURRENT behaviour so any future alias layer is a
    deliberate, test-visible change — not a silent one."""

    def test_brukinsa_rename_splits_into_exit_and_entrant(self, oracle):
        k_old = _identity("ZANUBRUTINIB", "BRUKINSA", "BEIGENE SWITZ GMBH", "80MG")
        k_new = _identity("ZANUBRUTINIB", "BRUKINSA", "BEONE MED I GMBH", "80MG")
        # different normalised company → distinct identities
        assert k_old != k_new
        assert _norm_company("BEIGENE SWITZ GMBH") == "beigene switz"
        assert _norm_company("BEONE MED I GMBH") == "beone med i"
        # old-company key exits, new-company key enters
        assert k_old in oracle["exits"]
        assert k_new in oracle["entrants"]
        assert oracle["exits"][k_old]["dollars"] == 191_056_615
        assert oracle["entrants"][k_new]["dollars"] == 164_715_086


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║ D. Materiality gate — real moves vs suppressed drift, abs() symmetry           ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class TestMaterialityGate:
    def test_real_decrease_above_threshold_flagged(self, diff, oracle):
        """TARO-CLARITHROMYCIN 500MG: $411,752→$227,665 (Δ −44.7%), 3,298→1,836 units
        (Δ −44.3%). An equal-magnitude DECREASE must flag exactly like an increase
        (gate uses abs())."""
        k = _identity("CLARITHROMYCIN", "TARO-CLARITHROMYCIN", "SUN PHARMA CANADA", "500MG")
        assert k in oracle["moves"]
        r = _move_row(diff, "TARO-CLARITHROMYCIN", "500MG")
        assert int(r["Dollars Old"]) == 411_752 and int(r["Dollars New"]) == 227_665
        assert int(r["Dollars Δ"]) == -184_087
        assert int(r["Units Δ"]) == -1_462
        assert float(r["Dollars Δ%"]) == -44.7

    def test_abs_symmetry_of_gate(self):
        base = {"dollars": 1_000_000, "units": 50_000, "ext_units": 0}
        up = {"dollars": 1_300_000, "units": 50_000, "ext_units": 0}    # +30% / +$300k
        down = {"dollars": 700_000, "units": 50_000, "ext_units": 0}    # -30% / -$300k
        assert _material(base, up) is True
        assert _material(base, down) is True  # equal magnitude, opposite sign

    def test_sub_threshold_drift_suppressed(self, oracle):
        """A small rolling-MAT drift on a large base is NOT a move (both floors fail)."""
        # INVOKAMET 1000MG/150MG: +$8,917 (0.3%) / +98 units — far below both floors.
        # Locate via the oracle (brand + strength set) rather than guessing molecule.
        inv = [key for key in oracle["unchanged"]
               if key[1] == "invokamet" and key[3] == ("1000MG", "150MG")]
        assert inv, "INVOKAMET unchanged key not found"
        for key in inv:
            o, n = oracle["unchanged"][key]
            assert not _material(o, n)

    def test_units_abs_cleared_but_pct_floor_suppresses(self, oracle):
        """PREMARIN 0.625MG: units Δ +2,336 clears the 1,000-unit ABS floor, but at
        0.59% it fails the 10% PCT floor → suppressed. Proves the gate needs BOTH."""
        prem = [key for key in oracle["unchanged"] if key[1] == "premarin" and key[3] == ("0.625MG",)]
        assert prem, "PREMARIN unchanged key not found"
        for key in prem:
            o, n = oracle["unchanged"][key]
            assert abs(n["units"] - o["units"]) >= IQVIA_DIFF_UNITS_ABS      # abs floor cleared
            assert abs(n["units"] - o["units"]) / o["units"] < IQVIA_DIFF_PCT  # pct floor fails
            assert not _material(o, n)


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║ D. Core compare + edge cases — sheet selection, parsing, derived edge inputs  ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

def _subset_xlsx(raw: pd.DataFrame, molecules: list[str]) -> bytes:
    """Serialise a molecule-filtered slice of a real raw frame back to xlsx bytes
    (single 'data' sheet). Drops the hidden _excel_row helper column. Values are the
    real metric integers — nothing invented."""
    sub = raw[raw["Combined Molecule"].astype(str).str.strip().isin(molecules)]
    sub = sub[[c for c in sub.columns if not str(c).startswith("_")]]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        sub.to_excel(xw, sheet_name="data", index=False)
    return buf.getvalue()


def _subset_csv(raw: pd.DataFrame, molecules: list[str], thousands: bool = False) -> bytes:
    """Same slice as _subset_csv's xlsx twin, but CSV. With thousands=True the metric
    cells get comma group separators and zeros are written as '-' (the real CSV's
    formatting) so the parser's separator/blank handling is exercised."""
    sub = raw[raw["Combined Molecule"].astype(str).str.strip().isin(molecules)]
    sub = sub[[c for c in sub.columns if not str(c).startswith("_")]].copy()
    metric = detect_metric_columns(sub)
    if thousands:
        for c in metric:
            sub[c] = sub[c].map(lambda v: "-" if int(v) == 0 else f"{int(v):,}")
    buf = io.StringIO()
    sub.to_csv(buf, index=False)
    return buf.getvalue().encode("utf-8")


class TestSheetSelectionAndParsing:
    def test_pivot_sheet_skipped_data_chosen(self, real_bytes):
        _, new_b = real_bytes
        xls = pd.ExcelFile(io.BytesIO(new_b))
        assert "Pivot" in xls.sheet_names and "data" in xls.sheet_names
        assert _pick_data_sheet(xls) == "data"

    def test_date_label_formats_resolve_per_file(self, real_bytes):
        old_b, new_b = real_bytes
        old_cols = detect_metric_columns(parse_iqvia(old_b))
        new_cols = detect_metric_columns(parse_iqvia(new_b))
        # old file writes YYYY/MM, new file writes MM/YYYY — both resolve to (Y, M).
        assert any(c == "Dollars MAT 2025/06" for c in old_cols)
        assert any(c == "Dollars MAT 12/2025" for c in new_cols)
        assert latest_mat_metrics(old_cols)[0] == OLD_PERIOD
        assert latest_mat_metrics(new_cols)[0] == NEW_PERIOD
        assert _parse_mat_period("Dollars MAT 2025/06") == _parse_mat_period("Dollars MAT 06/2025") == OLD_PERIOD

    def test_csv_and_xlsx_parse_to_identical_values(self, raw_frames):
        """A real subset written as xlsx (numeric) and as CSV (comma thousands + '-'
        blanks) must collapse to identical metric values."""
        _, new_raw = raw_frames
        mols = ["AFLIBERCEPT", "NIRSEVIMAB", "CLOBETASOL"]
        x = collapse_iqvia(parse_iqvia(_subset_xlsx(new_raw, mols)))
        c = collapse_iqvia(parse_iqvia(_subset_csv(new_raw, mols, thousands=True)))
        metric = detect_metric_columns(x)
        x = x.sort_values(_ID).reset_index(drop=True)
        c = c.sort_values(_ID).reset_index(drop=True)
        assert list(x[_ID].itertuples(index=False)) == list(c[_ID].itertuples(index=False))
        for col in metric:
            assert x[col].tolist() == c[col].tolist(), f"value mismatch in {col}"

    def test_collapse_yields_one_row_per_4field_group(self, raw_frames):
        """collapse_iqvia guarantees a unique key per (mol, product, mfr, strength) —
        so the diff's set difference can never double-count."""
        _, new_raw = raw_frames
        col = collapse_iqvia(new_raw)
        assert not col.duplicated(subset=_ID).any()


class TestDerivedEdgeCases:
    def test_identical_files_zero_everything(self, raw_frames):
        _, new_raw = raw_frames
        b = _subset_xlsx(new_raw, ["AFLIBERCEPT", "CLOBETASOL"])
        d = compare_iqvia(b, b)
        assert len(d.entrants) == 0 and len(d.exits) == 0 and len(d.moves) == 0

    def test_fully_disjoint_files_only_entrants_and_exits(self, raw_frames):
        _, new_raw = raw_frames
        old_b = _subset_xlsx(new_raw, ["AFLIBERCEPT", "NIRSEVIMAB"])
        new_b = _subset_xlsx(new_raw, ["CLOBETASOL", "CLARITHROMYCIN"])
        d = compare_iqvia(old_b, new_b)
        assert len(d.moves) == 0
        assert len(d.entrants) > 0 and len(d.exits) > 0
        # nothing shared → no key appears in both signal tabs
        assert _df_keyset(d.entrants) & _df_keyset(d.exits) == set()

    def test_empty_extract_one_side(self, raw_frames):
        """An extract with the real header but zero data rows: the other file's rows
        all become exits (it was the OLD slot); no crash, no moves/entrants."""
        _, new_raw = raw_frames
        full = _subset_xlsx(new_raw, ["AFLIBERCEPT"])
        empty = _subset_xlsx(new_raw, ["__none__"])  # filter matches nothing → header only
        d = compare_iqvia(full, empty)
        assert len(d.entrants) == 0 and len(d.moves) == 0
        assert len(d.exits) > 0

    def test_single_row_extract(self, raw_frames):
        """A single real raw row (with non-zero latest sales) parses, collapses, and
        compares without error → exactly one exit."""
        _, new_raw = raw_frames
        _, latest = latest_mat_metrics(detect_metric_columns(new_raw))
        afl = new_raw[new_raw["Combined Molecule"].astype(str).str.strip() == "AFLIBERCEPT"]
        one = afl[afl[latest["dollars"]].astype("int64") > 0].head(1)
        assert len(one) == 1
        one = one[[c for c in one.columns if not str(c).startswith("_")]]
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as xw:
            one.to_excel(xw, sheet_name="data", index=False)
        single = buf.getvalue()
        empty = _subset_xlsx(new_raw, ["__none__"])
        d = compare_iqvia(single, empty)   # the single row is the only old product
        assert len(d.exits) == 1
        assert len(d.entrants) == 0 and len(d.moves) == 0
