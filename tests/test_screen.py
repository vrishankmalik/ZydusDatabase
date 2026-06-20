"""Tests for app/enrichment/screen.py — go/no-go product screening.

Fully offline: a synthetic built-workbook dataset (Sheet 1 + Sheet 2 GSUR) with
KNOWN competitor / filing / approval counts and a GOLDEN IQVIA sum is screened
against a fixed criteria set; the qualifying products and the exact six summary
values are asserted.

Golden IQVIA anchors reuse the human-verified MAT 12/2025 values from
tests/test_iqvia.py (the progesterone sample), so the value/quantity sums are
pinned to real, previously-computed numbers:
    02516187  SANIS / PROGESTERONE       Units 218,591  Dollars 21,215,081
    02493578  AURO  / AURO-PROGESTERONE  Units 233,159  Dollars 13,005,865
"""
from __future__ import annotations

import io

import openpyxl
import pandas as pd

from app.enrichment.screen import (
    apply_criteria,
    build_filtered_workbook,
    compute_products,
    latest_metric_columns,
    parse_criteria,
)

# Golden per-DIN IQVIA values (din -> (dollars, units, ext_units)).
_G_SANIS = (21215081, 218591, 100)
_G_AURO = (13005865, 233159, 200)

# Latest-period metric column names (mirror the real IQVIA header form).
_DOLLARS = "Dollars MAT 12/2025"
_UNITS = "Units MAT 12/2025"
_EXT = "Ext Units MAT 12/2025"

# Product-group golden aggregates for (PROGESTERONE, Capsule).
_GOLD_VALUE = _G_SANIS[0] + _G_AURO[0]   # 34,220,946
_GOLD_UNITS = _G_SANIS[1] + _G_AURO[1]   # 451,750
_GOLD_EXT = _G_SANIS[2] + _G_AURO[2]     # 300


def _sheet1() -> pd.DataFrame:
    """One PROGESTERONE 'Capsule' product (3 DINs) + a verbatim-distinct
    'Capsule (extended-release)' product (1 DIN)."""
    return pd.DataFrame([
        # Capsule product — two MARKETED competitors, one CANCELLED sibling.
        {"din": "02516187", "ingredient": "PROGESTERONE", "dosage_form": "Capsule",
         "company": "SANIS HEALTH INC", "status": "marketed",
         _DOLLARS: _G_SANIS[0], _UNITS: _G_SANIS[1], _EXT: _G_SANIS[2]},
        {"din": "02493578", "ingredient": "PROGESTERONE", "dosage_form": "Capsule",
         "company": "AURO PHARMA INC", "status": "marketed",
         _DOLLARS: _G_AURO[0], _UNITS: _G_AURO[1], _EXT: _G_AURO[2]},
        # Cancelled sibling with NO IQVIA match → contributes 0 to sums, not a
        # competitor, but still a distinct company that holds an approval.
        {"din": "09999999", "ingredient": "PROGESTERONE", "dosage_form": "Capsule",
         "company": "OLD PHARMA LTD", "status": "cancelled post market",
         _DOLLARS: None, _UNITS: None, _EXT: None},
        # Verbatim-distinct dosage form → its OWN product, never merged with Capsule.
        {"din": "02505223", "ingredient": "PROGESTERONE",
         "dosage_form": "Capsule (extended-release)",
         "company": "BIJUVA CO", "status": "marketed",
         _DOLLARS: 500, _UNITS: 5, _EXT: 1},
    ])


def _sheet2() -> pd.DataFrame:
    """GSUR filings: 3 progesterone submissions + 1 unrelated."""
    return pd.DataFrame([
        {"medicinal_ingredient": "progesterone", "company": "GenA"},
        {"medicinal_ingredient": "progesterone", "company": "GenB"},
        {"medicinal_ingredient": "progesterone", "company": "Not available"},
        {"medicinal_ingredient": "metformin", "company": "GenC"},
    ])


# ── latest-period detection ───────────────────────────────────────────────────

def test_latest_metric_columns_picks_newest_period():
    cols = [
        "din", "ingredient",
        "Dollars MAT 12/2024", "Units MAT 12/2024", "Ext Units MAT 12/2024",
        "Dollars MAT 12/2025", "Units MAT 12/2025", "Ext Units MAT 12/2025",
    ]
    latest = latest_metric_columns(cols)
    assert latest["value"] == "Dollars MAT 12/2025"
    assert latest["quantity"] == "Units MAT 12/2025"
    assert latest["quantity_ext"] == "Ext Units MAT 12/2025"


def test_latest_metric_columns_absent_when_no_iqvia():
    latest = latest_metric_columns(["din", "ingredient", "company", "status"])
    assert latest == {"value": None, "quantity": None, "quantity_ext": None}


# ── product computation (all six criteria) ────────────────────────────────────

def test_compute_products_all_six_exact_values():
    products, warnings = compute_products(_sheet1(), _sheet2())
    assert warnings == []  # no blank dosage forms

    cap = products[
        (products["ingredient"] == "PROGESTERONE")
        & (products["dosage_form"] == "Capsule")
    ]
    assert len(cap) == 1, "Expected exactly one (PROGESTERONE, Capsule) product"
    row = cap.iloc[0]

    # 1. competitors = distinct MARKETED companies (cancelled sibling excluded)
    assert row["competitors"] == 2
    # 2. filings = GSUR submissions matching the ingredient
    assert row["filings"] == 3
    # 3. approvals = distinct companies across ALL the product's DINs
    assert row["approvals"] == 3
    # 4-6. golden IQVIA sums (unmatched DIN contributes 0)
    assert row["value_sizeable"] == _GOLD_VALUE
    assert row["quantity_sizeable"] == _GOLD_UNITS
    assert row["quantity_ext_sizeable"] == _GOLD_EXT


def test_verbatim_dosage_form_is_a_separate_product():
    products, _ = compute_products(_sheet1(), _sheet2())
    forms = set(products["dosage_form"])
    assert "Capsule" in forms
    assert "Capsule (extended-release)" in forms, (
        "release-type modifier must never be collapsed into the base form"
    )
    er = products[products["dosage_form"] == "Capsule (extended-release)"].iloc[0]
    assert er["competitors"] == 1
    assert er["value_sizeable"] == 500


def test_strength_is_not_part_of_product_key():
    """Manufacturer view: different strengths of the same molecule + form are ONE
    product; a different dosage form is a different product. Strength never splits."""
    s1 = pd.DataFrame([
        # DPD embeds the strength inside the ingredient string ("METFORMIN 200 MG").
        {"din": "1", "ingredient": "METFORMIN 200 MG", "dosage_form": "Tablet",
         "company": "A", "status": "marketed"},
        {"din": "2", "ingredient": "METFORMIN 400 MG", "dosage_form": "Tablet",
         "company": "B", "status": "marketed"},
        {"din": "3", "ingredient": "METFORMIN 500 MG", "dosage_form": "Tablet",
         "company": "A", "status": "marketed"},
        # Same molecule, DIFFERENT form → a separate product.
        {"din": "4", "ingredient": "METFORMIN 200 MG", "dosage_form": "Solution",
         "company": "C", "status": "marketed"},
    ])
    products, _ = compute_products(s1, pd.DataFrame())
    pairs = set(zip(products["ingredient"], products["dosage_form"]))
    assert pairs == {("METFORMIN", "Tablet"), ("METFORMIN", "Solution")}

    tab = products[products["dosage_form"] == "Tablet"].iloc[0]
    # Three strengths collapsed into one Tablet product: companies A, B (A deduped).
    assert tab["competitors"] == 2
    assert tab["approvals"] == 2
    assert sorted(tab["_dins"]) == ["1", "2", "3"]


def test_combo_ingredient_strength_stripped_and_deduped():
    """A titration kit 'ALPELISIB 50 MG; ALPELISIB 200 MG' is mono-ALPELISIB after
    stripping strengths; its 'Tablet; Kit' form keeps it distinct from plain Tablet."""
    s1 = pd.DataFrame([
        {"din": "1", "ingredient": "ALPELISIB 150 MG", "dosage_form": "Tablet",
         "company": "NOVARTIS", "status": "marketed"},
        {"din": "2", "ingredient": "ALPELISIB 50 MG; ALPELISIB 200 MG",
         "dosage_form": "Tablet; Kit", "company": "NOVARTIS", "status": "marketed"},
    ])
    products, _ = compute_products(s1, pd.DataFrame())
    pairs = set(zip(products["ingredient"], products["dosage_form"]))
    assert pairs == {("ALPELISIB", "Tablet"), ("ALPELISIB", "Tablet; Kit")}


def test_unmatched_din_contributes_zero_not_excluded():
    """The cancelled DIN with no IQVIA values adds 0 — it neither inflates nor
    voids the Capsule product's golden sum."""
    products, _ = compute_products(_sheet1(), _sheet2())
    cap = products[products["dosage_form"] == "Capsule"].iloc[0]
    assert cap["value_sizeable"] == _GOLD_VALUE  # exactly the two matched DINs


# ── filtering: exact qualifying set ───────────────────────────────────────────

def test_filter_returns_exactly_expected_products():
    products, _ = compute_products(_sheet1(), _sheet2())
    # competitors above 1 AND value above 30,000,000 → only the Capsule product
    # (34.2M, 2 competitors) qualifies; the ER form (500, 1 competitor) does not.
    criteria = parse_criteria([
        {"metric": "competitors", "operator": "above", "value": 1},
        {"metric": "value", "operator": "above", "value": 30_000_000},
    ])
    qualifying = apply_criteria(products, criteria)
    pairs = set(zip(qualifying["ingredient"], qualifying["dosage_form"]))
    assert pairs == {("PROGESTERONE", "Capsule")}


def test_filter_below_and_exactly_operators():
    products, _ = compute_products(_sheet1(), _sheet2())
    # exactly 1 competitor → only the ER form
    q = apply_criteria(products, parse_criteria(
        [{"metric": "competitors", "operator": "exactly", "value": 1}]))
    assert set(q["dosage_form"]) == {"Capsule (extended-release)"}
    # below 2 approvals → only the ER form (1 approval)
    q2 = apply_criteria(products, parse_criteria(
        [{"metric": "approvals", "operator": "below", "value": 2}]))
    assert set(q2["dosage_form"]) == {"Capsule (extended-release)"}


def test_no_criteria_returns_all_products():
    products, _ = compute_products(_sheet1(), _sheet2())
    q = apply_criteria(products, [])
    assert len(q) == len(products) == 2


# ── workbook assembly ─────────────────────────────────────────────────────────

def test_build_filtered_workbook_two_tabs_and_summary_values():
    criteria = parse_criteria([
        {"metric": "competitors", "operator": "above", "value": 1},
        {"metric": "value", "operator": "above", "value": 30_000_000},
    ])
    xlsx, summary, detail, warnings = build_filtered_workbook(
        _sheet1(), _sheet2(), criteria
    )
    assert warnings == []

    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert wb.sheetnames == ["Summary", "Detail"]

    # Summary: one qualifying product row, all six values exact.
    sws = wb["Summary"]
    headers = [c.value for c in sws[1]]
    assert headers == [
        "Ingredient", "Dosage Form", "Number of Competitors",
        "Number of Filings", "Number of Approvals", "Value Sizeable ($)",
        "Quantity Sizeable (Units)", "Quantity Ext Sizeable",
    ]
    assert sws.max_row == 2, "exactly one qualifying product"
    vals = {headers[i]: sws.cell(2, i + 1).value for i in range(len(headers))}
    assert vals["Ingredient"] == "PROGESTERONE"
    assert vals["Dosage Form"] == "Capsule"
    assert vals["Number of Competitors"] == 2
    assert vals["Number of Filings"] == 3
    assert vals["Number of Approvals"] == 3
    assert vals["Value Sizeable ($)"] == _GOLD_VALUE
    assert vals["Quantity Sizeable (Units)"] == _GOLD_UNITS
    assert vals["Quantity Ext Sizeable"] == _GOLD_EXT

    # Detail: every DIN of the qualifying Capsule product (the 3 Capsule DINs),
    # not the ER form's DIN.
    dws = wb["Detail"]
    dheaders = [c.value for c in dws[1]]
    din_col = dheaders.index("DIN") + 1
    detail_dins = {dws.cell(r, din_col).value for r in range(2, dws.max_row + 1)}
    assert detail_dins == {"02516187", "02493578", "09999999"}


def test_value_criterion_without_iqvia_raises():
    """A value/quantity criterion with no IQVIA metric columns must fail loud."""
    s1 = pd.DataFrame([
        {"din": "1", "ingredient": "PROGESTERONE", "dosage_form": "Capsule",
         "company": "A", "status": "marketed"},
    ])
    criteria = parse_criteria([{"metric": "value", "operator": "above", "value": 1}])
    try:
        build_filtered_workbook(s1, pd.DataFrame(), criteria)
    except ValueError as exc:
        assert "IQVIA" in str(exc)
    else:
        raise AssertionError("Expected ValueError when value criterion lacks IQVIA data")


def test_blank_dosage_form_warns_but_keeps_product():
    s1 = pd.DataFrame([
        {"din": "1", "ingredient": "PROGESTERONE", "dosage_form": None,
         "company": "A", "status": "marketed"},
    ])
    products, warnings = compute_products(s1, pd.DataFrame())
    assert "PROGESTERONE" in warnings, "blank dosage form must raise a warning"
    # The product is still present (own group), not silently dropped.
    assert (products["ingredient"] == "PROGESTERONE").any()


def test_parse_criteria_skips_blank_and_unset_rows():
    raw = [
        {"metric": "competitors", "operator": "above", "value": 3},
        {"metric": "filings", "operator": "below", "value": ""},   # blank → skip
        {"metric": "", "operator": "above", "value": 1},            # no metric → skip
    ]
    crits = parse_criteria(raw)
    assert len(crits) == 1
    assert crits[0].metric == "competitors"
