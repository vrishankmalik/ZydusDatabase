"""Two-tab enriched workbook builder.

Sheet 1 — "DPD + NOC + Patents":
  One row per DIN (DPD and NOC records joined), sorted ascending by DIN.
  NOC rows whose DIN is blank / "Not Applicable" are excluded entirely.
  Patent block: WIDE format — patent_1_number/filing/grant/expiry through
  patent_M_number/..., where M = max patents held by any DIN in the dataset.
  Labeling columns per DIN (from store). Data protection columns from the
  Register of Innovative Drugs (dp_6yr_no_file_date, pediatric_extension,
  data_protection_ends).

  Columns removed vs old format (Change 2):
    - All *_url columns (record_url, noc_record_url, labeling_pdf_url)
    - All *_page citation columns
    - Old combined patent_numbers / all_patents_detail / earliest_* / latest_* columns
  Columns kept:
    - _drug_code (internal DPD identifier, not a URL)
    - needs_ocr (live OCR-pipeline provenance flag)

Sheet 2 — "Generic Submissions":
  GSUR records filtered to the queried ingredient (same normalisation used
  elsewhere). Standalone — never joined to Sheet 1.

CLI:
  python -m app.enrichment.workbook --q "alpelisib" --field ingredient
"""
from __future__ import annotations

import datetime
import io
import logging
import re
from typing import Any, Optional

import pandas as pd

from app.config import WORKBOOK_MIN_FILL_RATE

logger = logging.getLogger(__name__)

from app.enrichment.data_protection import (
    _match_data_protection_deterministic,
)
from app.enrichment.store import get_labeling_for_din, get_patents_for_din
from app.models import DrugRecord, SearchResponse

# DIN values that should be excluded from Sheet 1
_EXCLUDED_DIN_VALUES = {"", "not applicable", "n/a", "na", "none"}

# Supplement submission types to drop from Sheet 1 (SNDS / SANDS) — existing behavior preserved
_SUPPLEMENT_TYPE_RE = re.compile(
    r"\bSNDS\b|\bSANDS\b|Supplement\s+to\s+(a\s+New|an\s+Abbreviated)",
    re.IGNORECASE,
)

# ANDS lineage detector — used to partition NDS history from ANDS entries.
_ANDS_RE = re.compile(
    r"\bANDS\b|Abbreviated\s+New\s+Drug\s+Submission",
    re.IGNORECASE,
)

# Separator for multi-entry NOC history cells (newline is safe; API fields are single-line).
_NOC_JOIN_SEP = "\n"

# Sentinel dict for DPD DINs that have no matching NOC record
_NO_NOC_RECORD = {
    "noc_date": "No NOC record",
    "reason_for_supplement": "No NOC record",
    "submission_class": "No NOC record",
    "noc_submission_type": "No NOC record",
    "noc_therapeutic_class": "No NOC record",
}

# Sentinel for DINs present in NOC but absent from DPD (no product entry found)
_NO_DPD_RECORD = {
    "brand_name": "Not in DPD",
    "company": "Not in DPD",
    "ingredient": "Not in DPD",
    "strength": "Not in DPD",
    "dosage_form": "Not in DPD",
    "route": "Not in DPD",
    "status": "Not in DPD",
    "_drug_code": None,
    "_schedule": None,
}

_LABELING_FIELDS = (
    "active_ingredient", "nonmedicinal_ingredients",
    "pack_size", "pack_style",
    "color", "shape", "size_mm", "weight", "ph",
)

# Columns that are NEVER pruned regardless of fill rate.
_NEVER_DROP_COLS = frozenset({
    # Identity / provenance
    "ingredient_name", "din", "_drug_code",
    # DPD core
    "brand_name", "company", "ingredient", "strength",
    "dosage_form", "route", "status",
    # Patent summary
    "patent_count", "patent_number", "patent_grant_date", "patent_expiry_date",
    # NOC
    "noc_date", "reason_for_supplement", "submission_class",
    "noc_submission_type", "noc_therapeutic_class",
    # Labeling
    "active_ingredient", "nonmedicinal_ingredients",
    "pack_size", "pack_style",
    "color", "shape", "size_mm", "weight", "ph",
    # Data protection (always present even when no record matches)
    "dp_6yr_no_file_date", "pediatric_extension", "data_protection_ends",
})

# Canonical Sheet 1 column order — fixed, no dynamic patent groups.
# DIN is first so every row is immediately identifiable.
_SHEET1_COLS = (
    "din",
    "ingredient_name",
    "dosage_form",
    "ingredient", "brand_name", "company", "strength",
    "route", "status", "_drug_code", "_schedule",
    "noc_date", "reason_for_supplement", "submission_class",
    "noc_submission_type", "noc_therapeutic_class",
    "patent_count", "patent_number", "patent_grant_date", "patent_expiry_date",
    "active_ingredient", "nonmedicinal_ingredients",
    "pack_size", "pack_style", "color", "shape", "size_mm",
    "weight", "ph",
    "dp_6yr_no_file_date", "pediatric_extension", "data_protection_ends",
)

# Display-name overrides: internal key → header shown in the XLSX.
# All other column keys are title-cased via _col_to_header().
_HEADER_DISPLAY: dict[str, str] = {
    "din": "DIN",
    "ingredient": "SKU Name",
}


def _col_to_header(col_name: str) -> str:
    """Return the display header for a column key."""
    if col_name in _HEADER_DISPLAY:
        return _HEADER_DISPLAY[col_name]
    return col_name.replace("_", " ").title()


def _apply_display_names(df: pd.DataFrame) -> pd.DataFrame:
    """Rename columns to their display names for XLSX output (single-product path)."""
    rename = {k: v for k, v in _HEADER_DISPLAY.items() if k in df.columns}
    return df.rename(columns=rename) if rename else df


def _is_empty_for_fill(v: Any) -> bool:
    """True if v counts as empty for fill-rate purposes.

    None, NaN, "", and whitespace-only strings are empty.
    Sentinel strings ("No NOC record", "Not in PM", "No", …) are NOT empty —
    they represent a real, meaningful absence and protect the column.
    """
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    return str(v).strip() == ""


def _col_fill_rate(series: "pd.Series[Any]", n_rows: int) -> float:
    return int(series.apply(lambda v: not _is_empty_for_fill(v)).sum()) / n_rows


def _prune_sparse_columns(
    df: pd.DataFrame,
    min_fill_rate: float = WORKBOOK_MIN_FILL_RATE,
) -> pd.DataFrame:
    """Drop Sheet 1 columns whose non-empty fill rate is at or below min_fill_rate."""
    if df.empty:
        return df

    n_rows = len(df)
    cols_before = len(df.columns)
    cols_to_drop: list[str] = []
    report_lines: list[str] = []

    for col in df.columns:
        if col in _NEVER_DROP_COLS:
            continue
        fr = _col_fill_rate(df[col], n_rows)
        if fr <= min_fill_rate:
            n_filled = round(fr * n_rows)
            cols_to_drop.append(col)
            report_lines.append(
                f"  {col}: {n_filled}/{n_rows} = {fr:.1%} fill → dropped"
            )

    cols_after = cols_before - len(cols_to_drop)
    print(f"\n=== Workbook column cleanup (min_fill_rate={min_fill_rate:.1%}) ===")
    if report_lines:
        for line in report_lines:
            print(line)
    else:
        print("  (no columns dropped)")
    print(f"  Columns: {cols_before} → {cols_after}")
    print("=" * 52)

    return df.drop(columns=cols_to_drop)


# ── Sheet 1 helpers ───────────────────────────────────────────────────────────

# Matches a bare mass strength that has no volume denominator yet:
# e.g. "25 MG", "0.5 MCG", "100 IU" — but not "25 MG/ML" or "10 MG/5 ML".
_BARE_MASS_RE = re.compile(
    r"^([\d.,]+\s+(?:MG|MCG|G|IU|UNITS?|MEQ)\b)\s*$",
    re.IGNORECASE,
)


def _normalize_solution_strength(
    strength: Optional[str],
    dosage_form: Optional[str],
) -> Optional[str]:
    """For Solution dosage forms only: append /ML when strength is a bare mass.

    '25 MG' + form 'Solution' → '25 MG/ML'.
    Strengths that already contain '/' are left unchanged.
    Non-Solution forms are left unchanged.
    """
    if not strength or not dosage_form:
        return strength
    if "solution" not in dosage_form.lower():
        return strength
    if "/" in strength:
        return strength
    m = _BARE_MASS_RE.match(strength.strip())
    if m:
        return m.group(1).rstrip() + "/ML"
    return strength


def _is_excluded_din(din: Optional[str]) -> bool:
    return din is None or din.strip().lower() in _EXCLUDED_DIN_VALUES


def _collect_noc_no_din_records(records: list[DrugRecord]) -> list[DrugRecord]:
    """Return NOC records that carry no valid DIN (cannot produce a Sheet 1 row)."""
    return [r for r in records if r.source == "NOC" and _is_excluded_din(r.din)]


def _collect_dpd_rows(records: list[DrugRecord]) -> dict[str, dict[str, Any]]:
    """Build DIN-keyed dict from DPD records. record_url excluded per Change 2."""
    out: dict[str, dict[str, Any]] = {}
    for r in records:
        if r.source != "DPD" or _is_excluded_din(r.din):
            continue
        din = r.din.strip()  # type: ignore[union-attr]
        out[din] = {
            "din": din,
            "brand_name": r.brand_name,
            "company": r.company,
            "ingredient": r.ingredient,
            "strength": _normalize_solution_strength(r.strength, r.dosage_form),
            "dosage_form": r.dosage_form,
            "route": r.route,
            "status": r.status,
            "_drug_code": r.source_specific.get("drug_code"),
            "_schedule": r.source_specific.get("schedule"),
        }
    return out


def _collect_noc_rows(records: list[DrugRecord]) -> dict[str, dict[str, Any]]:
    """Build DIN-keyed dict from NOC records.

    SNDS/SANDS are dropped (existing behavior). Remaining entries are NDS or
    ANDS. NDS entries are accumulated chronologically ascending and joined with
    _NOC_JOIN_SEP so each cell holds the full NDS history. ANDS entries are
    kept only when no NDS entry exists for a DIN (using the most-recent ANDS).
    """
    # Accumulate per-DIN: separate NDS and ANDS entry lists.
    nds_entries: dict[str, list[dict[str, Any]]] = {}
    ands_entries: dict[str, dict[str, Any]] = {}

    for r in records:
        if r.source != "NOC" or _is_excluded_din(r.din):
            continue
        sub_type = r.source_specific.get("submission_type") or ""
        if _SUPPLEMENT_TYPE_RE.search(sub_type):
            continue  # drop SNDS / SANDS rows
        din = r.din.strip()  # type: ignore[union-attr]
        entry = {
            "noc_date": r.source_specific.get("noc_date"),
            "reason_for_supplement": r.source_specific.get("reason_for_supplement"),
            "submission_class": r.source_specific.get("submission_class"),
            "noc_submission_type": sub_type or None,
            "noc_therapeutic_class": r.source_specific.get("therapeutic_class"),
        }
        if _ANDS_RE.search(sub_type):
            # Keep only the most-recent ANDS (overwrite with later dates).
            prev = ands_entries.get(din)
            if prev is None or (entry["noc_date"] or "") >= (prev["noc_date"] or ""):
                ands_entries[din] = entry
        else:
            nds_entries.setdefault(din, []).append(entry)

    out: dict[str, dict[str, Any]] = {}

    # NDS DINs — sort ascending by noc_date and join multi-entry fields.
    for din, entries in nds_entries.items():
        entries.sort(key=lambda e: e["noc_date"] or "")
        def _join(field: str) -> Optional[str]:
            vals = [str(e[field]) if e[field] is not None else "" for e in entries]
            # Join ALL entries (one line per record); never skip blank values so
            # every column has the same line count and line N refers to record N.
            combined = _NOC_JOIN_SEP.join(vals)
            return combined or None
        out[din] = {
            "noc_date": _join("noc_date"),
            "reason_for_supplement": _join("reason_for_supplement"),
            "submission_class": _join("submission_class"),
            "noc_submission_type": _join("noc_submission_type"),
            "noc_therapeutic_class": _join("noc_therapeutic_class"),
        }

    # ANDS-only DINs — use the single most-recent ANDS entry.
    for din, entry in ands_entries.items():
        if din not in out:
            out[din] = entry

    return out


def _aggregate_patents_latest(
    din: str,
    as_of: Optional["datetime.date"] = None,
) -> dict[str, Any]:
    """Return active-patent summary: patent_count + patent_number/grant_date/expiry_date.

    as_of: cut-off date for "active" (expiry > as_of); defaults to today.

    patent_count = number of patents whose expiry_date is strictly after as_of.
    If stored patents exist but all are expired, patent_count is the sentinel
    string "all patents expired" and the detail cells are blank.
    If no patents are stored for this DIN, patent_count is 0 and details are None.

    The detail columns show the latest-expiry active patent (tiebreak: highest
    patent_number string).  They are blank when all patents are expired.
    """
    import datetime as _dt
    if as_of is None:
        as_of = _dt.date.today()

    rows = get_patents_for_din(din)

    def _parse_date(s: Any) -> "_dt.date":
        if not s:
            return _dt.date.min
        try:
            return _dt.date.fromisoformat(str(s)[:10])
        except (ValueError, TypeError):
            return _dt.date.min

    if not rows:
        return {
            "patent_count": 0,
            "patent_number": None,
            "patent_grant_date": None,
            "patent_expiry_date": None,
        }

    active = [r for r in rows if _parse_date(r.get("expiry_date")) > as_of]
    if not active:
        return {
            "patent_count": "all patents expired",
            "patent_number": None,
            "patent_grant_date": None,
            "patent_expiry_date": None,
        }

    best = max(active, key=lambda r: (_parse_date(r.get("expiry_date")), r.get("patent_number") or ""))
    return {
        "patent_count": len(active),
        "patent_number": best.get("patent_number"),
        "patent_grant_date": best.get("grant_date"),
        "patent_expiry_date": best.get("expiry_date"),
    }


def _get_labeling_cols(din: str) -> dict[str, Any]:
    """Return labeling fields for a DIN.

    *_page citation columns, labeling_pdf_url, and needs_ocr are excluded
    from the workbook output per schema requirements.
    """
    row = get_labeling_for_din(din)
    out: dict[str, Any] = {}
    for field in _LABELING_FIELDS:
        out[field] = row.get(field) if row else None
    return out


def _get_dp_cols(
    dpd_ingredient: Optional[str],
    dpd_company: Optional[str],
    dp_table: Optional[list[dict]],
) -> dict[str, Any]:
    """Return data protection fields for a DIN, or blanks when dp_table is None."""
    blank = {"dp_6yr_no_file_date": None, "pediatric_extension": None, "data_protection_ends": None}
    if dp_table is None:
        return blank
    matched = _match_data_protection_deterministic(
        dpd_ingredient or "", dpd_company or "", dp_table
    )
    if matched:
        return matched
    return blank


def _col_is_all_empty(series: pd.Series) -> bool:
    """True iff every value in the series is None/NaN/empty string/whitespace.

    Sentinel strings ("No NOC record", "No PM available", "Not in PM", …) are
    NOT empty — they carry real information and prevent the column from dropping.
    """
    for val in series:
        if val is None:
            continue
        try:
            if pd.isna(val):
                continue
        except (TypeError, ValueError):
            pass
        s = str(val).strip()
        if not s or s.lower() in ("none", "nan"):
            continue
        # Has at least one real value (including any sentinel string)
        return False
    return True


def _drop_empty_sheet1_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Drop all-empty columns from Sheet 1 before writing.

    Protected columns (_NEVER_DROP_COLS) are always kept even when all-empty.
    "Empty" means None/NaN/''/whitespace only; sentinels count as REAL data.
    """
    if df.empty:
        return df

    cols_to_drop: list[str] = []
    for col in df.columns:
        if col in _NEVER_DROP_COLS:
            continue
        if _col_is_all_empty(df[col]):
            cols_to_drop.append(col)

    if cols_to_drop:
        logger.info(
            "Dropping %d all-empty columns from Sheet 1: %s",
            len(cols_to_drop), cols_to_drop,
        )
        print(f"[workbook] Dropping {len(cols_to_drop)} all-empty Sheet 1 columns: {cols_to_drop}")

    return df.drop(columns=cols_to_drop)


def build_sheet1(
    response: SearchResponse,
    dp_table: Optional[list[dict]] = None,
    ingredient_name: Optional[str] = None,
    as_of: Optional["datetime.date"] = None,
) -> pd.DataFrame:
    """Build Sheet 1: one row per DIN, DPD + NOC + patents + labeling + data protection.

    A DIN appears only when it is present in BOTH NOC and DPD. NOC-only DINs (in
    NOC but absent from DPD) are excluded — they carry no DPD product data and are
    not useful in the export (e.g. DINs 02272113, 02272121). DPD-only DINs (in DPD
    but absent from NOC) are likewise excluded; they appear in build_exclusion_list().
    """
    all_records = [r for s in response.sources for r in s.records]

    dpd_by_din = _collect_dpd_rows(all_records)
    noc_by_din = _collect_noc_rows(all_records)

    # Log NOC entries that have no attached DIN — they cannot produce a row.
    for r in _collect_noc_no_din_records(all_records):
        logger.warning(
            "NOC entry has no DIN — excluded from sheet: brand=%s noc_date=%s",
            r.brand_name, r.source_specific.get("noc_date"),
        )

    # A DIN must exist in BOTH NOC and DPD to appear in Sheet 1.
    noc_dins = set(noc_by_din)
    dpd_dins = set(dpd_by_din)
    dpd_only = sorted(dpd_dins - noc_dins)
    noc_only = sorted(noc_dins - dpd_dins)
    if dpd_only:
        logger.info(
            "Excluding %d DPD-only DIN(s) not in NOC for %r: %s",
            len(dpd_only), response.metadata.query, dpd_only,
        )
    if noc_only:
        logger.info(
            "Excluding %d NOC-only DIN(s) not in DPD for %r: %s",
            len(noc_only), response.metadata.query, noc_only,
        )

    all_dins = sorted(noc_dins & dpd_dins)
    if not all_dins:
        logger.warning(
            "No DINs present in both NOC and DPD for %r — Sheet 1 will be empty.",
            response.metadata.query,
        )
        return pd.DataFrame()

    rows = []
    for din in all_dins:
        row: dict[str, Any] = {"din": din}
        dpd_rec = dpd_by_din.get(din)
        row.update(dpd_rec if dpd_rec is not None else _NO_DPD_RECORD)
        noc_data = noc_by_din.get(din)
        row.update(noc_data if noc_data is not None else _NO_NOC_RECORD)
        row.update(_aggregate_patents_latest(din, as_of=as_of))
        if dpd_rec is not None:
            row.update(_get_labeling_cols(din))
        else:
            # No DPD record → no drug_code → PM was never fetched; mark all labeling fields.
            row.update({field: "Not in DPD" for field in _LABELING_FIELDS})
        row.update(_get_dp_cols(dpd_rec.get("ingredient") if dpd_rec else None,
                                dpd_rec.get("company") if dpd_rec else None, dp_table))
        rows.append(row)

    df = pd.DataFrame(rows)

    # Apply the canonical fixed column order; append any unexpected extras at end.
    present = set(df.columns)
    ordered = [c for c in _SHEET1_COLS if c in present]
    ordered += [c for c in df.columns if c not in set(ordered)]

    df = df[ordered].sort_values("din", kind="stable").reset_index(drop=True)
    # ingredient_name column is in _SHEET1_COLS at position 0; populate it here.
    df["ingredient_name"] = ingredient_name or response.metadata.query
    # Re-apply ordering so ingredient_name stays first after the assignment.
    ordered2 = [c for c in _SHEET1_COLS if c in set(df.columns)]
    ordered2 += [c for c in df.columns if c not in set(ordered2)]
    df = df[ordered2]
    df = _drop_empty_sheet1_cols(df)
    return _prune_sparse_columns(df)


def build_exclusion_list(
    response: SearchResponse,
    ingredient_name: Optional[str] = None,
) -> pd.DataFrame:
    """Return a DataFrame of DPD DINs excluded because they are not in NOC.

    Columns: din, brand_name, company, ingredient, reason.
    This is the sidecar companion to build_sheet1 — every DIN that DPD returned
    but that is absent from NOC appears here with an explanation.
    """
    all_records = [r for s in response.sources for r in s.records]
    dpd_by_din = _collect_dpd_rows(all_records)
    noc_by_din = _collect_noc_rows(all_records)

    query = ingredient_name or response.metadata.query
    excluded_dins = sorted(set(dpd_by_din) - set(noc_by_din))
    rows = [
        {
            "din": din,
            "brand_name": dpd_by_din[din].get("brand_name"),
            "company": dpd_by_din[din].get("company"),
            "ingredient": dpd_by_din[din].get("ingredient"),
            "reason": f"not present in NOC for {query}",
        }
        for din in excluded_dins
    ]
    return pd.DataFrame(rows, columns=["din", "brand_name", "company", "ingredient", "reason"])


# ── Sheet 2 helpers ───────────────────────────────────────────────────────────

def _ingredient_matches(record_ingredient: Optional[str], query: str) -> bool:
    """Return True if query is contained in the record's ingredient string."""
    if not record_ingredient:
        return False
    q = re.sub(r"\s+", " ", query.strip()).lower()
    ing = re.sub(r"\s+", " ", record_ingredient.strip()).lower()
    return q in ing


def build_sheet2(
    response: SearchResponse,
    ingredient_name: Optional[str] = None,
) -> pd.DataFrame:
    """Build Sheet 2: Generic Submissions filtered to the queried ingredient."""
    query = response.metadata.query
    rows = []
    for sr in response.sources:
        if sr.source != "GenericSubmissions":
            continue
        for r in sr.records:
            if not _ingredient_matches(r.ingredient, query):
                continue
            rows.append({
                "medicinal_ingredient": r.ingredient,
                "company": r.company,
                "therapeutic_area": r.source_specific.get("therapeutic_area"),
                "year_month_accepted": r.source_specific.get("date_accepted"),
                "status": r.status,
            })

    label = ingredient_name or query
    if not rows:
        return pd.DataFrame(
            columns=["ingredient_name", "medicinal_ingredient", "company",
                     "therapeutic_area", "year_month_accepted", "status"]
        )
    df = pd.DataFrame(rows)
    df.insert(0, "ingredient_name", label)
    return df


# ── Workbook assembly ─────────────────────────────────────────────────────────

def _style_sheet(worksheet: Any, df: pd.DataFrame) -> None:
    """Apply bold header, freeze row, autofilter, and autosized columns."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.fill = header_fill
        cell.alignment = _center

    worksheet.freeze_panes = "A2"

    if not df.empty:
        last_col = get_column_letter(len(df.columns))
        worksheet.auto_filter.ref = f"A1:{last_col}1"

    for i, col in enumerate(df.columns, 1):
        max_val_len = (
            df[col].fillna("").astype(str).str.len().max()
            if not df.empty else 0
        )
        width = min(max(len(str(col)) + 2, int(max_val_len or 0) + 2), 60)
        worksheet.column_dimensions[get_column_letter(i)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _build_status_sheet(
    writer: Any,
    response: SearchResponse,
    source_errors: dict[str, Optional[str]],
) -> None:
    """Write a per-source status warning sheet (used when allow_partial=True)."""
    rows = []
    for src in response.sources:
        rows.append({
            "source": src.source,
            "status": src.status,
            "record_count": src.count,
            "error_message": src.error_message or "",
            "warning": (
                "⚠ DATA MISSING FROM THIS EXPORT"
                if src.status == "error"
                else ""
            ),
        })
    df = pd.DataFrame(
        rows,
        columns=["source", "status", "record_count", "error_message", "warning"],
    )
    sheet_name = "⚠ Source Status"
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    _style_sheet(writer.sheets[sheet_name], df)


def build_workbook(
    response: SearchResponse,
    source_errors: Optional[dict[str, Optional[str]]] = None,
    dp_table: Optional[list[dict]] = None,
    as_of: Optional[datetime.date] = None,
) -> bytes:
    """Assemble the enriched workbook and return XLSX bytes.

    source_errors: when provided (allow_partial=True path), appends a
    '⚠ Source Status' sheet that visibly flags every failed source.
    dp_table: pre-fetched active data protection register rows (from
    fetch_data_protection_table()); None means the three dp_* columns are blank.
    as_of: reference date for patent activity (default = today).
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet1 = build_sheet1(response, dp_table=dp_table, as_of=as_of)
        sheet1_out = _apply_display_names(sheet1)
        sheet1_out.to_excel(writer, sheet_name="DPD + NOC + Patents", index=False)
        _style_sheet(writer.sheets["DPD + NOC + Patents"], sheet1_out)

        sheet2 = build_sheet2(response)
        sheet2.to_excel(writer, sheet_name="Generic Submissions", index=False)
        _style_sheet(writer.sheets["Generic Submissions"], sheet2)

        if source_errors is not None:
            _build_status_sheet(writer, response, source_errors)

    return buf.getvalue()


# ── Multi-product side-by-side workbook ──────────────────────────────────────

# Light-fill palette for product blocks.  8 distinct accessible colors;
# cycles for 9+ products.  All are light enough that black text remains
# readable; distinct enough that adjacent products are easy to tell apart.
_BLOCK_COLORS: list[str] = [
    "EDD6EB",  # light purple   (#AA55A0 ~15%)
    "CCF0F0",  # light teal     (#00A5A5 ~20%)
    "DDD5EE",  # light deep purple (#3D226E ~20%)
    "CCE7F2",  # light teal-dark   (#008BAD ~20%)
    "F3E5F2",  # pale purple
    "D5F2F2",  # pale teal
    "E5D8F0",  # lavender purple
    "C5EAEA",  # seafoam teal
]

# Medium tints of Zydus purple/teal for banner rows (kept for potential future use)
_BLOCK_BANNER_COLORS: list[str] = [
    "D4A8D0",  # medium purple   (#AA55A0 ~50%)
    "80CECE",  # medium teal     (#00A5A5 ~50%)
    "9D8AC4",  # medium deep purple (#3D226E ~50%)
    "7FC4D8",  # medium teal-dark   (#008BAD ~50%)
    "E0B8DC",  # light-medium purple
    "99D8D8",  # light-medium teal
    "C4AADA",  # lavender purple
    "88CCCC",  # teal variant
]


def _block_color(idx: int) -> str:
    return _BLOCK_COLORS[idx % len(_BLOCK_COLORS)]


def _block_banner_color(idx: int) -> str:
    return _BLOCK_BANNER_COLORS[idx % len(_BLOCK_BANNER_COLORS)]


def _safe_cell_val(val: Any) -> Any:
    """Convert a pandas/numpy value to a plain Python type for openpyxl."""
    import numpy as np  # numpy is a pandas dependency; always available
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(val, np.integer):
        return int(val)
    if isinstance(val, np.floating):
        return float(val)
    if isinstance(val, np.bool_):
        return bool(val)
    return val


def _write_vertical_sheet(
    ws: Any,
    df: pd.DataFrame,
    ingredient_palette: list[tuple[str, str]],
) -> None:
    """Write a vertically-stacked multi-ingredient sheet.

    Layout:
      Row 1  — KEY LEGEND: "PRODUCT KEY:" + one colored cell per ingredient.
      Row 2  — HEADERS: shared column names (purple fill, white text).
      Row 3+ — DATA: rows sorted by ingredient order then DIN, color-coded by
                     ingredient.  Even rows within each ingredient block get the
                     block fill; odd rows get white (zebra stripe).

    ingredient_palette: list of (name, block_hex_color) in display order.
    freeze_panes="A3" keeps the key legend and headers visible while scrolling.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    import math as _math

    KEY_ROW    = 1
    HEADER_ROW = 2
    DATA_START = 3

    HEADER_FILL    = PatternFill(start_color="3D226E", end_color="3D226E", fill_type="solid")
    KEY_LABEL_FILL = PatternFill(start_color="3D226E", end_color="3D226E", fill_type="solid")
    WHITE_FILL     = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    _thin = Side(style="thin", color="D1D1D1")
    _med  = Side(style="medium", color="A0A0A0")
    CELL_BORDER   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    HEADER_BORDER = Border(left=_thin, right=_thin, top=_med,  bottom=_med)

    _STATUS_COLOR = {
        "marketed":  "1A6B3C",
        "approved":  "1A6B3C",
        "cancelled": "9B1C1C",
        "dormant":   "7A4F00",
        "inactive":  "7A4F00",
    }
    _MIN_WIDTHS = {
        "din": 12, "status": 13, "form": 13, "route": 13,
        "strength": 14, "noc_date": 13, "patent_count": 10,
        "ingredient_name": 18,
    }

    # Build ingredient → block fill map
    color_map: dict[str, PatternFill] = {
        name: PatternFill(start_color=color, end_color=color, fill_type="solid")
        for name, color in ingredient_palette
    }

    n_cols = len(df.columns) if not df.empty else 0

    # ── Row 1: Key legend ─────────────────────────────────────────────────────
    label_cell = ws.cell(row=KEY_ROW, column=1)
    label_cell.value = "INGREDIENT KEY:"
    label_cell.font = Font(bold=True, name="Calibri", size=10, color="FFFFFF")
    label_cell.fill = KEY_LABEL_FILL
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = CELL_BORDER

    for i, (name, color) in enumerate(ingredient_palette):
        key_col = 2 + i
        cell = ws.cell(row=KEY_ROW, column=key_col)
        cell.value = name
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(key_col)].width = max(len(name) + 4, 16)

    ws.row_dimensions[KEY_ROW].height = 22

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    cols = list(df.columns) if not df.empty else []
    for j, col_name in enumerate(cols, 1):
        cell = ws.cell(row=HEADER_ROW, column=j)
        cell.value = _col_to_header(col_name)
        cell.font = Font(bold=True, name="Calibri", size=10, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = HEADER_BORDER
    ws.row_dimensions[HEADER_ROW].height = 26

    if not df.empty:
        last_col = get_column_letter(n_cols)
        ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{HEADER_ROW}"

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths: list[float] = []
    for j, col_name in enumerate(cols, 1):
        max_val_len = (
            df[col_name].fillna("").astype(str).str.len().max()
            if not df.empty else 0
        )
        floor = _MIN_WIDTHS.get(col_name.lower(), 0)
        width = min(max(len(str(col_name)) + 4, int(max_val_len or 0) + 3, floor), 42)
        col_widths.append(width)
        ws.column_dimensions[get_column_letter(j)].width = width

    # ── Rows 3+: Data ─────────────────────────────────────────────────────────
    # Track per-ingredient row counter for zebra striping within each block.
    ing_row_counter: dict[str, int] = {}

    LINE_HEIGHT_PT = 15.0

    for r_idx, (_idx, row_series) in enumerate(df.iterrows()):
        ing = str(row_series.get("ingredient_name", "")) if "ingredient_name" in df.columns else ""
        ing_row_counter[ing] = ing_row_counter.get(ing, 0)
        row_fill = (
            color_map.get(ing, WHITE_FILL)
            if ing_row_counter[ing] % 2 == 0
            else WHITE_FILL
        )
        ing_row_counter[ing] += 1

        excel_row = DATA_START + r_idx
        max_lines = 1

        for j, col_name in enumerate(cols):
            val = _safe_cell_val(row_series[col_name])
            cell = ws.cell(row=excel_row, column=j + 1)
            cell.value = val
            cell.fill = row_fill
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            status_key = str(val).lower().strip() if val is not None else ""
            if col_name == "status" and status_key in _STATUS_COLOR:
                cell.font = Font(bold=True, name="Calibri", size=10,
                                 color=_STATUS_COLOR[status_key])
            else:
                cell.font = Font(name="Calibri", size=10)
            text_len = len(str(val)) if val is not None else 0
            col_w = col_widths[j] if j < len(col_widths) else 20
            lines = _math.ceil(text_len / max(col_w, 1)) if text_len else 1
            max_lines = max(max_lines, lines)

        ws.row_dimensions[excel_row].height = max(LINE_HEIGHT_PT, max_lines * LINE_HEIGHT_PT)

    # ── Freeze key legend + header rows ──────────────────────────────────────
    ws.freeze_panes = "A3"


def _write_reconciliation_sheet(wb: Any, recon_df: pd.DataFrame) -> None:
    """Write the IQVIA Reconciliation sheet to the workbook."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title="IQVIA Reconciliation")

    STATUS_FILLS = {
        "matched":           "C6EFCE",  # green
        "ambiguous":         "FFEB9C",  # yellow
        "low_score":         "FFD7B5",  # orange
        "no_din_match":      "FFC7CE",  # red
        "din_no_iqvia_match": "E2EFDA",  # light green (DIN present, just no IQVIA data)
    }

    cols = list(recon_df.columns)
    header_fill = PatternFill(start_color="3D226E", end_color="3D226E", fill_type="solid")

    for j, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j)
        cell.value = col.replace("_", " ").title()
        cell.font = Font(bold=True, name="Calibri", size=10, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_len = max(len(col) + 4, 14)
        ws.column_dimensions[get_column_letter(j)].width = min(col_len, 40)

    status_col_idx = cols.index("status") + 1 if "status" in cols else None

    for r_idx, (_, row) in enumerate(recon_df.iterrows(), 2):
        status = str(row.get("status", "")) if "status" in row else ""
        hex_fill = STATUS_FILLS.get(status, "FFFFFF")
        row_fill = PatternFill(start_color=hex_fill, end_color=hex_fill, fill_type="solid")

        for j, col in enumerate(cols, 1):
            val = _safe_cell_val(row[col])
            cell = ws.cell(row=r_idx, column=j)
            cell.value = val
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    if recon_df.shape[0] > 0:
        from openpyxl.utils import get_column_letter as gcl
        ws.auto_filter.ref = f"A1:{gcl(len(cols))}1"


def build_workbook_multiproduct(
    products: list[tuple[str, SearchResponse]],
    source_errors: Optional[dict[str, Optional[str]]] = None,
    dp_table: Optional[list[dict]] = None,
    iqvia_df: Optional["pd.DataFrame"] = None,
    debug_iqvia_rows: bool = False,
    as_of: Optional[datetime.date] = None,
) -> "tuple[bytes, pd.DataFrame, pd.DataFrame, pd.DataFrame]":
    """Build a vertically-stacked multi-ingredient two-tab workbook.

    Each product becomes one color-coded vertical block in the shared column
    layout on both tabs.  Rows for different ingredients are distinguished by
    their row fill color; the ``ingredient_name`` column (always first) makes
    the ingredient explicit in every row.

    Single-product is a degenerate case (one block) and produces the same
    underlying data as ``build_workbook``.

    iqvia_df: collapsed IQVIA DataFrame (output of collapse_iqvia()).  When
    provided, metric columns are appended to Sheet 1 by DIN matching, and an
    "IQVIA Reconciliation" sheet is added.

    Returns (xlsx_bytes, combined_sheet1_df, combined_sheet2_df, reconciliation_df).
    """
    import openpyxl

    colors = [_block_color(i) for i in range(len(products))]

    # ── Build per-product DataFrames (ingredient_name already inserted) ───────
    sheet1_frames: list[pd.DataFrame] = []
    sheet2_frames: list[pd.DataFrame] = []
    for (name, response), _color in zip(products, colors):
        s1 = build_sheet1(response, dp_table=dp_table, ingredient_name=name, as_of=as_of)
        s2 = build_sheet2(response, ingredient_name=name)
        sheet1_frames.append(s1)
        sheet2_frames.append(s2)

    # ── Vertically concatenate; pd.concat aligns columns by name ─────────────
    def _vstack(frames: list[pd.DataFrame]) -> pd.DataFrame:
        non_empty = [f for f in frames if not f.empty]
        if not non_empty:
            return pd.DataFrame()
        return pd.concat(non_empty, ignore_index=True, sort=False)

    combined_s1 = _vstack(sheet1_frames)
    combined_s2 = _vstack(sheet2_frames)

    # Restore canonical column order after concat (columns may shift)
    def _reorder(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        present = set(df.columns)
        ordered = [c for c in _SHEET1_COLS if c in present]
        ordered += [c for c in df.columns if c not in set(_SHEET1_COLS)]
        return df[ordered]

    combined_s1 = _reorder(combined_s1)

    # ── Attach IQVIA metrics (optional) ──────────────────────────────────────
    recon_df: pd.DataFrame = pd.DataFrame()
    if iqvia_df is not None and not iqvia_df.empty and not combined_s1.empty:
        from app.enrichment.iqvia import match_iqvia_to_sheet1
        combined_s1, recon_df = match_iqvia_to_sheet1(
            combined_s1, iqvia_df, debug_iqvia_rows=debug_iqvia_rows
        )

    # ── Ingredient palette (name → color) for the sheet writer ───────────────
    ingredient_palette = [(name, colors[i]) for i, (name, _) in enumerate(products)]

    # ── Assemble XLSX ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "DPD + NOC + Patents"
    _write_vertical_sheet(ws1, combined_s1, ingredient_palette)

    ws2 = wb.create_sheet(title="Generic Submissions")
    _write_vertical_sheet(ws2, combined_s2, ingredient_palette)

    if source_errors is not None:
        _build_status_sheet_multi(wb, products, source_errors)

    wb.save(buf)
    return buf.getvalue(), combined_s1, combined_s2, recon_df


def _build_status_sheet_multi(
    wb: Any,
    products: list[tuple[str, SearchResponse]],
    source_errors: dict[str, Optional[str]],
) -> None:
    """Append a ⚠ Source Status sheet when allow_partial=True."""
    from openpyxl.styles import Alignment, Font, PatternFill
    _center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws = wb.create_sheet(title="⚠ Source Status")
    headers = ["product", "source", "status", "record_count", "error_message", "warning"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = _center

    row_idx = 2
    for name, response in products:
        for src in response.sources:
            ws.cell(row=row_idx, column=1).value = name
            ws.cell(row=row_idx, column=2).value = src.source
            ws.cell(row=row_idx, column=3).value = src.status
            ws.cell(row=row_idx, column=4).value = src.count
            ws.cell(row=row_idx, column=5).value = src.error_message or ""
            ws.cell(row=row_idx, column=6).value = (
                "⚠ DATA MISSING FROM THIS EXPORT" if src.status == "error" else ""
            )
            for col_idx in range(1, 7):
                ws.cell(row=row_idx, column=col_idx).alignment = _center
            row_idx += 1


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    import asyncio
    import os

    parser = argparse.ArgumentParser(description="Build the two-tab enriched workbook.")
    parser.add_argument("--q", required=True, help="Search query")
    parser.add_argument("--field", default="ingredient",
                        help="ingredient | brand | company | din")
    parser.add_argument("--out", default=None, help="Output file path (.xlsx)")
    _args = parser.parse_args()

    async def _run() -> None:
        from app.sources.dpd import search_dpd
        from app.sources.generic_submissions import search_generic_submissions
        from app.sources.noc import search_noc
        from app.sources.patent_register import search_patent_register
        from app.normalize import normalize_query
        from app.models import SearchMetadata
        from app.enrichment.data_protection import fetch_data_protection_table
        from datetime import datetime, timezone

        canonical, extra_terms = await normalize_query(_args.q, _args.field)
        sources = await asyncio.gather(
            search_dpd(canonical, _args.field, extra_terms),
            search_generic_submissions(canonical, _args.field, extra_terms),
            search_noc(canonical, _args.field, extra_terms),
            search_patent_register(canonical, _args.field, extra_terms),
        )
        response = SearchResponse(
            metadata=SearchMetadata(
                query=_args.q,
                field=_args.field,
                timestamp=datetime.now(timezone.utc).isoformat(),
                normalized_terms=[canonical] + extra_terms,
            ),
            sources=list(sources),
        )

        # Enrich patents
        from app.enrichment.patents import enrich_patents
        all_dins = [
            r.din for s in response.sources for r in s.records
            if r.din and not _is_excluded_din(r.din)
        ]
        if all_dins:
            await enrich_patents(all_dins)

        dp_table = await fetch_data_protection_table()
        xlsx = build_workbook(response, dp_table=dp_table)
        out_path = _args.out or f"enriched_{_args.q.replace(' ', '_')}_{_args.field}.xlsx"
        with open(out_path, "wb") as fh:
            fh.write(xlsx)
        print(f"Wrote {os.path.getsize(out_path):,} bytes → {out_path}")

    asyncio.run(_run())
