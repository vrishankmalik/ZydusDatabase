"""Full DPD universe (no-PDF) loader, cache, and sheet builder.

This module powers the "Full universe" tab (options 3 & 4).  It is ADDITIVE and
reuses the existing building blocks read-only:

  * build_sheet1(..., include_dpd_only=True)  — union DIN universe (keeps the
    grandfathered / pre-NOC DPD-only products instead of dropping them).
  * match_iqvia_to_sheet1                       — DPD-native IQVIA sizing match.
  * build_filtered_workbook / compute_products  — the six-criteria screen.
  * enrich_labeling_batch_fast                  — survivor-only PM PDF enrichment
    (driven by the job runner in app/universe_job.py, never from here).

The cheap no-PDF universe is built from a SINGLE allfiles.zip pull (~1.4 MB,
~13.5k drug_codes, all status sets) rather than ~100k REST calls.  NOC, patents,
and data-protection columns attach blank-when-absent for DPD-only rows.

Freshness: the parsed universe is cached in-process for CACHE_TTL (4 h, mirrors
the HTTP cache).  Within that window option 4 reuses option 3's build — no double
pull.  After the window the next request re-downloads allfiles.zip and rebuilds.
reset_universe_cache() (wired into /api/reset-all-caches) drops the in-process
cache AND the on-disk extract so the Reset button forces a fresh DPD pull.
"""
from __future__ import annotations

import asyncio
import csv
import io
import logging
import shutil
import time
import zipfile
from pathlib import Path
from typing import Any, Optional

import pandas as pd

from app.config import CACHE_DIR, CACHE_TTL
from app.enrichment.store import get_labeling_for_din
from app.enrichment.workbook import (
    _LABELING_FIELDS,
    _apply_display_names,
    build_sheet1,
)
from app.models import DrugRecord, SearchMetadata, SearchResponse, SourceResult

logger = logging.getLogger(__name__)

# ── allfiles.zip source + on-disk extract location ────────────────────────────
DPD_ALLFILES_URL = (
    "https://www.canada.ca/content/dam/hc-sc/documents/services/"
    "drug-product-database/allfiles.zip"
)
UNIVERSE_CACHE_DIR = Path(CACHE_DIR) / "universe"

# 4-hour freshness window (mirrors CACHE_TTL = 14400).
UNIVERSE_TTL = CACHE_TTL

# Score band: a fuzzy IQVIA match at or above this is "high" confidence, below it
# "low" (the house-brand / generic-label collisions worth auditing at scale).
# CONFIDENT_THRESHOLD (65) is the matcher's floor for stamping a value at all.
LOW_CONFIDENCE_SCORE = 85.0

# ── Bulk extract column positions (0-indexed; Health Canada "what's in the
# extract" layout — verified against a known DIN at parse time in _verify) ─────
_DRUG_COL_CODE, _DRUG_COL_DIN, _DRUG_COL_BRAND = 0, 3, 4
_INGRED_COL_CODE, _INGRED_COL_NAME, _INGRED_COL_STRENGTH, _INGRED_COL_UNIT = 0, 2, 4, 5
_FORM_COL_CODE, _FORM_COL_NAME = 0, 2
_ROUTE_COL_CODE, _ROUTE_COL_NAME = 0, 2
_STATUS_COL_CODE, _STATUS_COL_FLAG, _STATUS_COL_STATUS = 0, 1, 2
_COMP_COL_CODE, _COMP_COL_NAME = 0, 3

_EXTRACT_FILES = ("drug.txt", "ingred.txt", "form.txt", "route.txt", "status.txt", "comp.txt")

# NOC sentinel emitted by build_sheet1 for a DIN with no NOC record; on the
# universe sheet we blank it (cite-or-blank: absent NOC → blank, not a label).
_NOC_SENTINEL = "No NOC record"
_NOC_COLS = (
    "noc_date", "reason_for_supplement", "submission_class",
    "noc_submission_type", "noc_therapeutic_class",
)


# ── In-process universe cache ─────────────────────────────────────────────────
def _build_dosage_form_map(dpd_records: list[DrugRecord]) -> dict[str, list[str]]:
    """Base dosage-form → raw-variants map across the whole DPD catalogue.

    Built from the parsed records' ``dosage_form`` cells using the SAME canonical
    collapse the filter uses, so the dropdown's base forms and the filter's
    matching never drift.  Powers the dosage-form dropdown on both tabs.
    """
    from app.enrichment.screen import build_dosage_form_map
    return build_dosage_form_map(r.dosage_form for r in dpd_records)


class UniverseBundle:
    """Parsed full-universe inputs (no PDF). Cached in-process for UNIVERSE_TTL.

    Also carries, on the SAME 4-hour freshness, the active data-protection register
    rows (``dp_table``) and the dosage-form base→raw map used by the filter UI's
    dropdown — both derived once per fresh build and reused for the cache window.
    """

    def __init__(
        self,
        dpd_records: list[DrugRecord],
        gsur_records: list[DrugRecord],
        dp_table: Optional[list[dict]] = None,
        dosage_form_map: Optional[dict[str, list[str]]] = None,
    ):
        self.dpd_records = dpd_records
        self.gsur_records = gsur_records
        self.dp_table = dp_table
        self.dosage_form_map = (
            dosage_form_map if dosage_form_map is not None
            else _build_dosage_form_map(dpd_records)
        )
        self.built_at = time.time()

    @property
    def age_seconds(self) -> float:
        return time.time() - self.built_at

    def is_fresh(self, ttl: int = UNIVERSE_TTL) -> bool:
        return self.age_seconds < ttl


_CACHE: dict[str, Optional[UniverseBundle]] = {"bundle": None}


def reset_universe_cache() -> int:
    """Drop the in-process universe AND the on-disk allfiles extract.

    Wired into /api/reset-all-caches so the existing Reset button forces the next
    full-universe request to re-download allfiles.zip from DPD.  Returns 1 if a
    cached bundle was present, else 0.
    """
    had = 1 if _CACHE.get("bundle") is not None else 0
    _CACHE["bundle"] = None
    try:
        if UNIVERSE_CACHE_DIR.exists():
            shutil.rmtree(UNIVERSE_CACHE_DIR)
    except OSError:
        logger.warning("Could not remove universe extract dir %s", UNIVERSE_CACHE_DIR)
    return had


def universe_cache_status() -> dict[str, Any]:
    """Snapshot of the cache state for the UI (cached?/age/expires_in)."""
    bundle = _CACHE.get("bundle")
    if bundle is None:
        return {"cached": False, "ttl_seconds": UNIVERSE_TTL}
    return {
        "cached": True,
        "ttl_seconds": UNIVERSE_TTL,
        "age_seconds": round(bundle.age_seconds, 1),
        "expires_in_seconds": round(max(0.0, UNIVERSE_TTL - bundle.age_seconds), 1),
        "fresh": bundle.is_fresh(),
        "dpd_records": len(bundle.dpd_records),
        "gsur_records": len(bundle.gsur_records),
    }


# ── allfiles.zip download + parse ─────────────────────────────────────────────
def _extract_is_fresh() -> bool:
    drug = UNIVERSE_CACHE_DIR / "drug.txt"
    if not drug.exists():
        return False
    return (time.time() - drug.stat().st_mtime) < UNIVERSE_TTL


def _download_extract() -> None:
    """Download allfiles.zip and extract the six files we parse (if stale)."""
    if _extract_is_fresh():
        return
    import httpx

    UNIVERSE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    logger.info("Downloading DPD allfiles.zip for full universe…")
    resp = httpx.get(DPD_ALLFILES_URL, follow_redirects=True, timeout=180.0)
    resp.raise_for_status()
    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        names = set(zf.namelist())
        for fname in _EXTRACT_FILES:
            if fname not in names:
                raise RuntimeError(
                    f"allfiles.zip is missing {fname!r} — extract layout changed."
                )
            (UNIVERSE_CACHE_DIR / fname).write_bytes(zf.read(fname))
    # Schema is verified by load_dpd_universe_records() (the single canonical parse
    # path) against whatever cache_dir it reads — including this one in production.


def _read_rows(path: Path) -> list[list[str]]:
    with open(path, encoding="latin-1", newline="") as fh:
        return list(csv.reader(fh))


def _verify_columns(cache_dir: Path) -> None:
    """Spot-check column positions against a known DIN (GLUCOPHAGE = 02099233).

    Fails loudly on schema drift rather than emitting silently-wrong universe data.
    """
    din = "02099233"
    code: Optional[str] = None
    for row in _read_rows(cache_dir / "drug.txt"):
        if len(row) > _DRUG_COL_DIN and row[_DRUG_COL_DIN] == din:
            code = row[_DRUG_COL_CODE]
            brand = row[_DRUG_COL_BRAND] if len(row) > _DRUG_COL_BRAND else ""
            assert "GLUCOPHAGE" in brand.upper(), f"drug.txt brand col drift: {brand!r}"
            break
    assert code is not None, f"Universe schema check: DIN {din} not in drug.txt"

    forms = {r[_FORM_COL_NAME].upper() for r in _read_rows(cache_dir / "form.txt")
             if r[_FORM_COL_CODE] == code and len(r) > _FORM_COL_NAME}
    assert any("TABLET" in f for f in forms), f"form.txt col drift for {code}: {forms!r}"

    comps = {r[_COMP_COL_NAME].strip() for r in _read_rows(cache_dir / "comp.txt")
             if r[_COMP_COL_CODE] == code and len(r) > _COMP_COL_NAME and r[_COMP_COL_NAME].strip()}
    assert comps, f"comp.txt company-name col drift for {code} (no company found)"


def _index_by_code(rows: list[list[str]], code_col: int) -> dict[str, list[list[str]]]:
    out: dict[str, list[list[str]]] = {}
    for r in rows:
        if len(r) > code_col and r[code_col]:
            out.setdefault(r[code_col], []).append(r)
    return out


def load_dpd_universe_records(cache_dir: Path = UNIVERSE_CACHE_DIR) -> list[DrugRecord]:
    """Parse an already-downloaded extract into DPD DrugRecords (one per drug_code
    with a DIN).

    Pure parse — it NEVER downloads.  ``get_universe`` calls ``_download_extract()``
    first to populate the production ``UNIVERSE_CACHE_DIR``; this function then reads
    whatever ``cache_dir`` it is given, so a caller (tests, fixtures) can parse a
    different extract dir without any network side-effect.  The column schema check
    runs here against ``cache_dir``, so offline fixture parsing is still guarded.

    Mirrors the field shapes app/sources/dpd.py produces from the REST API so the
    IQVIA matcher and the six-criteria screen behave identically on this layer.
    """
    _verify_columns(cache_dir)

    drug_rows = _read_rows(cache_dir / "drug.txt")
    ingred_by_code = _index_by_code(_read_rows(cache_dir / "ingred.txt"), _INGRED_COL_CODE)
    form_by_code = _index_by_code(_read_rows(cache_dir / "form.txt"), _FORM_COL_CODE)
    route_by_code = _index_by_code(_read_rows(cache_dir / "route.txt"), _ROUTE_COL_CODE)
    status_by_code = _index_by_code(_read_rows(cache_dir / "status.txt"), _STATUS_COL_CODE)
    comp_by_code = _index_by_code(_read_rows(cache_dir / "comp.txt"), _COMP_COL_CODE)

    def _col(row: list[str], idx: int) -> str:
        return row[idx].strip() if len(row) > idx else ""

    def _current_status(code: str) -> Optional[str]:
        rows = status_by_code.get(code, [])
        current = [r for r in rows if _col(r, _STATUS_COL_FLAG).upper().startswith("Y")]
        chosen = current or rows
        vals = [_col(r, _STATUS_COL_STATUS) for r in chosen if _col(r, _STATUS_COL_STATUS)]
        seen: set[str] = set()
        ordered = [v for v in vals if not (v in seen or seen.add(v))]
        return "; ".join(ordered) or None

    def _company(code: str) -> Optional[str]:
        for r in comp_by_code.get(code, []):
            name = _col(r, _COMP_COL_NAME)
            if name:
                return name
        return None

    records: list[DrugRecord] = []
    for drow in drug_rows:
        din = _col(drow, _DRUG_COL_DIN)
        code = _col(drow, _DRUG_COL_CODE)
        if not din or not code:
            continue
        brand = _col(drow, _DRUG_COL_BRAND) or None

        irows = sorted(
            ingred_by_code.get(code, []),
            key=lambda r: _col(r, _INGRED_COL_NAME).upper(),
        )
        ing_parts: list[str] = []
        strength_parts: list[str] = []
        all_names: list[str] = []
        for r in irows:
            name = _col(r, _INGRED_COL_NAME)
            if not name:
                continue
            all_names.append(name)
            s = _col(r, _INGRED_COL_STRENGTH)
            u = _col(r, _INGRED_COL_UNIT)
            su = f"{s} {u}".strip()
            ing_parts.append(f"{name} {su}".strip() if su else name)
            if s:
                strength_parts.append(su)
        ingredient_str = "; ".join(ing_parts) or None
        strength_str = "; ".join(strength_parts) or None

        forms = [_col(r, _FORM_COL_NAME) for r in form_by_code.get(code, []) if _col(r, _FORM_COL_NAME)]
        routes = [_col(r, _ROUTE_COL_NAME) for r in route_by_code.get(code, []) if _col(r, _ROUTE_COL_NAME)]

        records.append(DrugRecord(
            source="DPD",
            ingredient=ingredient_str,
            brand_name=brand,
            company=_company(code),
            din=din,
            all_ingredients=all_names,
            strength=strength_str,
            dosage_form="; ".join(forms) or None,
            route="; ".join(routes) or None,
            status=_current_status(code),
            source_specific={"drug_code": int(code) if code.isdigit() else code},
        ))
    logger.info("Parsed %d DPD universe records from allfiles.zip", len(records))
    return records


async def _load_gsur_records() -> list[DrugRecord]:
    """Fetch the FULL Generic Submissions table (all rows, unfiltered), read-only."""
    from app.sources.generic_submissions import _fetch_page, _parse_table, _RECORD_URL
    try:
        html = await _fetch_page()
        rows = _parse_table(html)
    except Exception:
        logger.warning("Full GSUR fetch failed for universe; GSUR sheet will be empty", exc_info=True)
        return []
    out: list[DrugRecord] = []
    for r in rows:
        ing = r["ingredient"]
        out.append(DrugRecord(
            source="GenericSubmissions",
            ingredient=ing,
            company=r["company"] if r["company"] != "Not available" else None,
            all_ingredients=(
                [i.strip() for i in ing.split(";") if i.strip()]
                if ";" in ing else ([ing.strip()] if ing.strip() else [])
            ),
            status="Under Review",
            record_url=_RECORD_URL,
            source_specific={
                "therapeutic_area": r["therapeutic_area"],
                "date_accepted": r["date_accepted"],
            },
        ))
    return out


async def get_universe(force_refresh: bool = False) -> UniverseBundle:
    """Return the cached universe bundle, rebuilding when stale or forced.

    Within UNIVERSE_TTL the same bundle is returned (option 4 reuses option 3's
    build).  A rebuild re-downloads allfiles.zip only when the on-disk extract is
    also stale (or was cleared by Reset).
    """
    bundle = _CACHE.get("bundle")
    if bundle is not None and bundle.is_fresh() and not force_refresh:
        return bundle

    # Ensure the production extract is present/fresh, then parse it.  Download and
    # parse are separate steps so load_dpd_universe_records() stays a pure, no-network
    # parser that honours its cache_dir argument.
    await asyncio.to_thread(_download_extract)
    dpd_records = await asyncio.to_thread(load_dpd_universe_records)
    gsur_records = await _load_gsur_records()
    # Active data-protection register, fetched once per fresh build and reused for
    # the 4-hour window (invalidated by reset-all-caches via reset_universe_cache).
    # fetch_data_protection_table() has its own 24-h HTTP cache, so this is cheap.
    from app.enrichment.data_protection import fetch_data_protection_table
    dp_table = await fetch_data_protection_table()
    bundle = UniverseBundle(dpd_records, gsur_records, dp_table=dp_table)
    _CACHE["bundle"] = bundle
    return bundle


# ── Universe response + sheet assembly ────────────────────────────────────────
def build_universe_response(bundle: UniverseBundle) -> SearchResponse:
    """Wrap the parsed bundle in a SearchResponse (DPD + GSUR sources, no NOC)."""
    from datetime import datetime, timezone
    return SearchResponse(
        metadata=SearchMetadata(
            query="Full DPD Universe",
            field="ingredient",
            timestamp=datetime.now(timezone.utc).isoformat(),
        ),
        sources=[
            SourceResult(source="DPD", status="ok", records=bundle.dpd_records),
            SourceResult(source="GenericSubmissions", status="ok", records=bundle.gsur_records),
        ],
    )


def build_universe_sheet2(bundle: UniverseBundle) -> pd.DataFrame:
    """Sheet 2 for the universe: ALL Generic Submissions rows (unfiltered)."""
    cols = ["ingredient_name", "medicinal_ingredient", "company",
            "therapeutic_area", "year_month_accepted", "status"]
    rows = [{
        "ingredient_name": "",
        "medicinal_ingredient": r.ingredient,
        "company": r.company,
        "therapeutic_area": r.source_specific.get("therapeutic_area"),
        "year_month_accepted": r.source_specific.get("date_accepted"),
        "status": r.status,
    } for r in bundle.gsur_records]
    return pd.DataFrame(rows, columns=cols)


def _blank_noc_sentinels(df: pd.DataFrame) -> pd.DataFrame:
    """Replace the 'No NOC record' sentinel with blank (cite-or-blank at row level).

    Runs AFTER build_sheet1 so the NOC columns have already survived column
    pruning (they are protected anyway via _NEVER_DROP_COLS).
    """
    for col in _NOC_COLS:
        if col in df.columns:
            df[col] = df[col].apply(lambda v: "" if v == _NOC_SENTINEL else v)
    return df


def _confidence_for(status: str, top_score: Optional[float], notes: str) -> str:
    """Map an IQVIA recon outcome to a per-DIN confidence label."""
    note = (notes or "").lower()
    if status == "matched":
        if note.startswith("exact-brand") or "generic-label alias" in note:
            return "exact"
        if top_score is not None and top_score >= LOW_CONFIDENCE_SCORE:
            return "high"
        return "low"
    if status == "low_score":
        return "low"
    return "none"


def attach_match_confidence(
    sheet1_df: pd.DataFrame,
    recon_df: pd.DataFrame,
) -> tuple[pd.DataFrame, int]:
    """Add an 'iqvia_match_confidence' column to Sheet 1; return (df, low_count).

    Derived ONLY from the matcher's reconciliation output — the IQVIA match logic
    itself is untouched.  At full-universe scale the fuzzy matches (house-brand /
    generic-label collisions: PRO DOC / JAMP / Pharmascience) are exactly the rows
    whose IQVIA sizing should be sanity-checked, so they are surfaced as 'low' and
    counted for the KPI.
    """
    df = sheet1_df.copy()
    if "din" not in df.columns:
        df["iqvia_match_confidence"] = ""
        return df, 0
    if recon_df is None or recon_df.empty:
        df["iqvia_match_confidence"] = ""
        return df, 0

    _RANK = {"matched": 3, "low_score": 2, "din_no_iqvia_match": 1}
    by_din: dict[str, tuple[int, str, Optional[float], str]] = {}
    for _, r in recon_df.iterrows():
        din = str(r.get("din", "") or "").strip()
        if not din:
            continue
        status = str(r.get("status", "") or "").strip()
        rank = _RANK.get(status, 0)
        ts = r.get("top_score")
        try:
            ts = float(ts) if ts is not None and str(ts) != "" else None
        except (TypeError, ValueError):
            ts = None
        prev = by_din.get(din)
        if prev is None or rank > prev[0]:
            by_din[din] = (rank, status, ts, str(r.get("notes", "") or ""))

    def _label(din_val: Any) -> str:
        din = str(din_val or "").strip()
        rec = by_din.get(din)
        if rec is None:
            return "none"
        _, status, ts, notes = rec
        return _confidence_for(status, ts, notes)

    df["iqvia_match_confidence"] = df["din"].apply(_label)
    low_count = int((df["iqvia_match_confidence"] == "low").sum())
    return df, low_count


def build_universe_sheet1(
    response: SearchResponse,
    iqvia_df: Optional[pd.DataFrame] = None,
    debug_iqvia_rows: bool = False,
    dp_table: Optional[list[dict]] = None,
) -> tuple[pd.DataFrame, pd.DataFrame, int]:
    """Build the no-PDF universe Sheet 1 (union DIN universe).

    Returns (sheet1_df, recon_df, low_confidence_count).  NOC is blank-when-absent;
    patents are blank for DPD-only rows.  Data protection is now populated via
    ``dp_table`` (the active Register of Innovative Drugs rows) using the same
    ingredient+manufacturer match as the per-product Search path; passing
    ``dp_table=None`` keeps the three dp_* columns blank (legacy behaviour).  IQVIA
    sizing + match confidence are attached when an IQVIA frame is supplied.
    """
    df = build_sheet1(response, dp_table=dp_table, ingredient_name="", include_dpd_only=True)
    df = _blank_noc_sentinels(df)

    recon_df = pd.DataFrame()
    if iqvia_df is not None and not iqvia_df.empty and not df.empty:
        from app.enrichment.iqvia import match_iqvia_to_sheet1
        df, recon_df = match_iqvia_to_sheet1(df, iqvia_df, debug_iqvia_rows=debug_iqvia_rows)

    df, low_count = attach_match_confidence(df, recon_df)
    return df, recon_df, low_count


# ── Survivor labeling patch (option 4) ────────────────────────────────────────
def patch_labeling_for_dins(sheet1_df: pd.DataFrame, dins: set[str]) -> pd.DataFrame:
    """Overwrite labeling columns for the given DINs from the labeling store.

    Used after survivor-only PM enrichment so only the filtered rows carry PDF
    fields; every other universe row keeps its blank labeling cells.
    """
    if sheet1_df.empty or not dins:
        return sheet1_df
    df = sheet1_df.copy()
    din_str = df["din"].astype(str).str.strip()
    cache: dict[str, dict] = {}
    for din in dins:
        row = get_labeling_for_din(din)
        if row:
            cache[din] = row
    for field in _LABELING_FIELDS:
        if field not in df.columns:
            continue
        df[field] = [
            (cache.get(d, {}).get(field) if d in cache else df.iloc[i][field])
            for i, d in enumerate(din_str)
        ]
    return df


# ── Disclaimer + workbook assembly (option 3) ─────────────────────────────────
UNIVERSE_DISCLAIMER_LINES = (
    "FULL DPD UNIVERSE. Product-Monograph (PDF) data is OMITTED.",
    "",
    "This sheet covers the entire Drug Product Database catalogue as of today, "
    "including legacy / grandfathered (pre-NOC) products. The per-product Product "
    "Monograph PDF fields (appearance, excipients, pH, pack size/style, etc.) are "
    "NOT included here because extracting them across the whole market would take "
    "hours.",
    "",
    "To pull those PDF fields, use 'Filter & enrich' on the Full universe tab: it "
    "applies your six-criteria filter first and then fetches + parses the Product "
    "Monographs only for the products that pass.",
    "",
    "Data-protection columns (six-year no-file date, pediatric extension, data "
    "protection ends) ARE populated here. The Register of Innovative Drugs is "
    "matched to every product by ingredient + manufacturer. The NOC and patent "
    "columns appear but are NOT populated. This no-PDF universe performs neither "
    "join, so every cell in them is blank by design (see the caution below). IQVIA "
    "sizing columns are present only when an IQVIA file was uploaded; check the "
    "'Iqvia Match Confidence' column. 'Low' flags fuzzy house-brand matches worth "
    "verifying before trusting the sizing.",
)

# Safety caution rendered as a bold red line on the Read-Me (matches the IQVIA
# low-confidence caution styling).  The remedy names ONLY the standard per-product
# export, which is the path that actually runs the NOC + patent + data-protection
# joins; 'Filter & enrich' (option 4) only adds Product-Monograph fields and still
# performs none of those three joins, so it must NOT be offered here.
UNIVERSE_NOT_EVALUATED_CAUTION = (
    "⚠ NOC and patent data are NOT evaluated in this view. A blank patent cell here "
    "means 'not assessed', NOT 'none'. Do NOT read it as 'no patent' or 'free to "
    "launch'. (Data protection IS evaluated here.) To assess NOC and patents for a "
    "product, run the standard per-product export (the Search tab → Export), which "
    "performs those joins."
)


# ── Streaming (write-only) workbook assembly ──────────────────────────────────
# The universe sheet is the whole DPD catalogue (~13.5k rows).  A normal openpyxl
# workbook holds every styled Cell object in RAM until save(), pushing the build
# request's peak RSS to ~530 MB — over the 512 MB free-tier container limit.
# openpyxl's write_only mode streams rows to a temp file via lxml and keeps almost
# no cell state, cutting the workbook-write overhead from ~130 MB to a few MB (so
# the build peaks ~220 MB).  Header rows are still styled (cheap); large data
# sheets stream unstyled values, which is the standard pattern for big exports.

def _wo_cell(ws: Any, value: Any, *, font=None, fill=None, alignment=None):
    """A styled WriteOnlyCell (the only way to style cells in write_only mode)."""
    from openpyxl.cell import WriteOnlyCell
    c = WriteOnlyCell(ws, value=value)
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if alignment is not None:
        c.alignment = alignment
    return c


def _wo_write_disclaimer_sheet(ws: Any, low_confidence_count: Optional[int]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    wrap = Alignment(wrap_text=True, vertical="top")
    body_font = Font(name="Calibri", size=10)
    red_font = Font(bold=True, name="Calibri", size=10, color="9B1C1C")

    ws.column_dimensions["A"].width = 110
    ws.append([_wo_cell(
        ws, "⚠ READ ME: PDF data omitted",
        font=Font(bold=True, size=12, color="FFFFFF", name="Calibri"),
        fill=PatternFill(start_color="C47F17", end_color="C47F17", fill_type="solid"),
    )])
    ws.append([])  # blank row 2 (matches the original layout)
    for line in UNIVERSE_DISCLAIMER_LINES:
        ws.append([_wo_cell(ws, line, font=body_font, alignment=wrap)])
    # Bold red safety caution — blank patent/NOC/DP cells must not be read as "none".
    ws.append([])
    ws.append([_wo_cell(ws, UNIVERSE_NOT_EVALUATED_CAUTION, font=red_font, alignment=wrap)])
    if low_confidence_count is not None:
        ws.append([])
        ws.append([_wo_cell(
            ws,
            f"IQVIA low-confidence (fuzzy) matches in this dataset: {low_confidence_count}",
            font=red_font,
        )])


def _wo_write_data_sheet(ws: Any, df: pd.DataFrame) -> None:
    """Stream a DataFrame to a write_only worksheet, byte-for-byte equivalent to the
    normal-mode ``_df_to_sheet`` + ``_style_sheet`` output.

    Reproduces the shared styling EXACTLY — grey bold centered header, frozen header
    row, header autofilter, content-aware column widths, and the per-data-cell
    Calibri-10 / centered / wrapped formatting.  The style objects are created once
    and reused (openpyxl deduplicates them), and write_only streams each row to the
    lxml temp file and releases it, so memory stays flat regardless of row count
    even though every cell is individually styled.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from app.enrichment.workbook import _safe_cell_val

    cols = list(df.columns)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True, name="Calibri", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    # Shared once; reused for every data cell (matches _style_sheet's per-cell style).
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # In write_only mode, sheet-level properties (column widths, freeze panes,
    # autofilter) must be set BEFORE the first append() or they are not written.
    for i, col in enumerate(cols, 1):
        max_val_len = (
            df[col].fillna("").astype(str).str.len().max() if not df.empty else 0
        )
        width = min(max(len(str(col)) + 2, int(max_val_len or 0) + 2), 60)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    if not df.empty:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    ws.append([
        _wo_cell(ws, str(col), font=header_font, fill=header_fill, alignment=header_align)
        for col in cols
    ])
    for row in df.itertuples(index=False, name=None):
        ws.append([
            _wo_cell(ws, _safe_cell_val(v), font=data_font, alignment=data_align)
            for v in row
        ])


def _wo_write_reconciliation_sheet(wb: Any, recon_df: pd.DataFrame) -> None:
    """Write-only twin of workbook._write_reconciliation_sheet (same colours)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from app.enrichment.workbook import _safe_cell_val

    ws = wb.create_sheet(title="IQVIA Reconciliation")
    STATUS_FILLS = {
        "matched": "C6EFCE", "ambiguous": "FFEB9C", "low_score": "FFD7B5",
        "no_din_match": "FFC7CE", "din_no_iqvia_match": "E2EFDA",
    }
    cols = list(recon_df.columns)
    header_fill = PatternFill(start_color="3D226E", end_color="3D226E", fill_type="solid")
    header_font = Font(bold=True, name="Calibri", size=10, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    body_font = Font(name="Calibri", size=10)
    top = Alignment(vertical="top")

    # Sheet-level props before the first append (write_only requirement).
    for j, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(max(len(col) + 4, 14), 40)
    ws.freeze_panes = "A2"
    if cols and recon_df.shape[0] > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    ws.append([
        _wo_cell(ws, col.replace("_", " ").title(), font=header_font, fill=header_fill, alignment=center)
        for col in cols
    ])
    fill_cache: dict[str, Any] = {}
    for row in recon_df.itertuples(index=False, name=None):
        rowmap = dict(zip(cols, row))
        status = str(rowmap.get("status", "") or "")
        hex_fill = STATUS_FILLS.get(status, "FFFFFF")
        fill = fill_cache.get(hex_fill)
        if fill is None:
            fill = PatternFill(start_color=hex_fill, end_color=hex_fill, fill_type="solid")
            fill_cache[hex_fill] = fill
        ws.append([
            _wo_cell(ws, _safe_cell_val(rowmap[c]), font=body_font, fill=fill, alignment=top)
            for c in cols
        ])


def build_universe_workbook(
    sheet1_df: pd.DataFrame,
    sheet2_df: pd.DataFrame,
    recon_df: Optional[pd.DataFrame] = None,
    low_confidence_count: Optional[int] = None,
) -> bytes:
    """Assemble the option-3 workbook: Disclaimer + universe data tabs (no PDF).

    Uses openpyxl write_only mode so a full ~13.5k-row universe builds within the
    512 MB free-tier memory limit (see the module-level note above).
    """
    import openpyxl

    buf = io.BytesIO()
    wb = openpyxl.Workbook(write_only=True)

    ws_disc = wb.create_sheet(title="⚠ Read Me")
    _wo_write_disclaimer_sheet(ws_disc, low_confidence_count)

    ws1 = wb.create_sheet(title="Full Universe (no PDF)")
    _wo_write_data_sheet(ws1, _apply_display_names(sheet1_df))

    ws2 = wb.create_sheet(title="Generic Submissions")
    _wo_write_data_sheet(ws2, sheet2_df)

    wb.save(buf)
    return buf.getvalue()
