"""Post-export IQVIA verifier — read-only, fully separate from the export path.

Why this exists
---------------
Manual spot-checking of the IQVIA numbers attached to DINs in a finished workbook
is finding a high error rate.  The errors are *match* errors, not arithmetic
errors: ``collapse_iqvia`` is a plain groupby-sum, so the numbers stamped onto a
DIN are always the exact sum of *some* IQVIA group — they are just sometimes the
*wrong* group's numbers.

The critical consequence: re-summing a DIN's source rows can only ever prove the
arithmetic.  It CANNOT catch a confident-wrong match, because the cited rows
belong to the wrong group and their sum still "agrees" with the cell.  So this
tool does not re-sum.  It independently judges whether each *match* is correct.

The dominant failure mode (confirmed against the real matcher)
--------------------------------------------------------------
``match_iqvia_to_sheet1`` scores a candidate as ``0.5*brand_sim + 0.5*company_sim``
and accepts at a combined threshold of 65.  When one manufacturer makes two
products at the same strength and only the *sibling* DIN is present in Sheet 1,
company similarity alone (≈100) drags the average past 65 even though the brand is
wrong — e.g. an IQVIA group "BRAND BETA / ACME" lands on a Sheet-1 DIN for
"BRAND ALPHA / ACME" at combined score 83 while ``brand_sim`` is only ~50.  The
match is flagged "matched", the sum reconciles, and it is still wrong.

The independent signal this tool uses is exactly the leg that the gate averaged
away: a correct match has a *strong brand* leg, not merely a strong company leg.

What it checks (independent of the accept gate)
-----------------------------------------------
For every DIN that carries IQVIA numbers in the workbook, the tool reverse-maps
the cell values back to the exact collapsed IQVIA group (the sum is exact ints),
then evaluates, reusing the project's own normalisation helpers:

  * company_carried  — the brand leg is weak (``brand_sim`` below BRAND_FLOOR)
                       while the company leg is strong: the match leaned on the
                       manufacturer.  This is the confident-wrong fingerprint.
  * better_din_for_group — a *different* Sheet-1 DIN at the same strength matches
                       this group's brand markedly better: the money may belong to
                       that DIN.
  * better_group_for_din — a *different* IQVIA group at the same strength matches
                       this DIN's brand markedly better: the DIN may be carrying
                       the wrong sibling's numbers.
  * value_orphan     — the workbook value matches no current IQVIA group at all
                       (stale workbook / wrong IQVIA file).
  * rerun_divergence — re-running the live matcher disagrees with the value in the
                       workbook (regression / stale workbook / plumbing).

The first three reuse only ``_norm_brand`` / ``_norm_company`` / ``_norm_strength``
/ ``_sim`` and the matcher's own exact strength-set prefilter; the last two re-run
``match_iqvia_to_sheet1`` unchanged.  Nothing here modifies, imports side effects
into, or adds any time to the export pipeline.

Output is a ranked list of suspicious DINs (most suspicious first) so the manual
"check 5, find 2" loop becomes one sweep.  Exit code is non-zero when anything is
flagged, so it can gate CI against a known-good fixture.

Usage
-----
    python -m app.enrichment.iqvia_verify --workbook export.xlsx --iqvia raw.xlsx
    python -m app.enrichment.iqvia_verify --workbook export.xlsx --iqvia raw.xlsx --json
"""
from __future__ import annotations

import argparse
import itertools
import re
import sys
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from app.enrichment.iqvia import (
    parse_iqvia,
    collapse_iqvia,
    detect_metric_columns,
    match_iqvia_to_sheet1,
    _norm_brand,
    _norm_company,
    _norm_strength,
    _sim,
    _is_generic_brand,
    _GENERIC_FILLER_WORDS,
    _METRIC_COL_RE,
    MIN_COMPANY_SIM,
)
from app.enrichment.workbook import _col_to_header

# ── Verifier thresholds (review heuristics — independent of the matcher gate) ──

# A match whose brand leg is below this is treated as "company-carried" and
# surfaced for review.  The matcher accepts at combined 65 with a 50/50 weight,
# so a match can clear the gate on company alone with a brand leg as low as ~30.
# Set above that floor: a genuine generic brand match (shared molecule suffix)
# normally scores well above 72.
BRAND_FLOOR = 72.0

# A rival (DIN or IQVIA group) must beat the assigned match's brand similarity by
# at least this margin before it is reported, to avoid noise from near-ties.
# EXCEPTION (recall): an *exact-brand counterpart* — a rival whose normalised brand
# similarity is a perfect 100 while the assigned match is below 100 — is reported
# regardless of this margin.  That is the precise fingerprint of the cross-company
# swap the matcher's exact-brand priority now fixes (e.g. PMS group landing on the
# PRZ DIN while the exact PMS DIN scores 100): the dollar-weighted miss whose margin
# (14.3) fell just under 15.  Once the matcher is correct this rule reports ZERO —
# that zero is itself the regression signal.
RIVAL_MARGIN = 15.0

# Normalised brand similarity that counts as an exact-brand counterpart.
EXACT_BRAND_SIM = 100.0

# A rival candidate is only worth surfacing if its own brand similarity is at
# least this high — otherwise "better" just means "less bad".
RIVAL_MIN_BRAND = 60.0

# The Sheet-1 internal column keys the matcher consumes.  Used both to rebuild a
# matcher-ready frame from the workbook and to build the display→key header map.
_SHEET1_KEYS = ("din", "ingredient", "brand_name", "company", "strength", "status")

# Severity weights for the composite suspicion score (higher = more suspicious).
_FLAG_WEIGHTS = {
    "value_orphan": 100.0,
    "rerun_divergence": 80.0,
    "company_carried": 1.0,      # multiplied by (BRAND_FLOOR - brand_sim)
    "better_din_for_group": 1.0,  # multiplied by (rival_brand - assigned_brand)
    "better_group_for_din": 1.0,  # multiplied by (rival_brand - assigned_brand)
}


@dataclass
class DinResult:
    din: str
    brand: str
    company: str
    strength: str
    values: dict[str, int]                 # metric col -> value carried in workbook
    assigned_product: Optional[str] = None  # IQVIA group the values came from
    assigned_manufacturer: Optional[str] = None
    assigned_strength: Optional[str] = None
    brand_sim: Optional[float] = None
    company_sim: Optional[float] = None
    flags: list[str] = field(default_factory=list)
    evidence: list[str] = field(default_factory=list)
    suspicion: float = 0.0
    # "high" = genuine review item; "low" = downranked false-positive class
    # (a generic/USP-style label that can only be the single same-company DIN, so
    # the weak brand leg is expected, not a wrong match). Low-priority items stay
    # in the report for audit but are kept out of the headline suspicious list.
    tier: str = "high"


@dataclass
class VerifyReport:
    dins: list[DinResult]                       # every DIN carrying IQVIA values
    suspicious: list[DinResult]                 # high-tier, suspicion > 0, ranked
    metric_cols: list[str]
    n_enriched: int
    iqvia_groups: int
    low_priority: list[DinResult] = field(default_factory=list)  # downranked tier

    @property
    def ok(self) -> bool:
        return not self.suspicious

    def get(self, din: str) -> Optional[DinResult]:
        for d in self.dins:
            if d.din == din:
                return d
        return None


# ── Workbook reading ──────────────────────────────────────────────────────────

_SHEET_NAME = "DPD + NOC + Patents"


def _build_header_map() -> dict[str, str]:
    """Display-header → internal-key map, reusing the project's own header fn.

    Accepts both layouts the exporter produces:
      * multi-product (vertical sheet): headers are ``_col_to_header(key)``
      * single-product (build_workbook): headers are the raw snake_case key,
        except din/ingredient which go through ``_apply_display_names``.
    """
    m: dict[str, str] = {}
    for key in _SHEET1_KEYS:
        m[_col_to_header(key)] = key   # title-cased / display form (multi-product)
        m[key] = key                   # raw snake_case form (single-product)
    return m


def _read_sheet1(workbook_path: str) -> pd.DataFrame:
    """Read the DPD/NOC/Patents sheet into a matcher-ready internal-keyed frame.

    Auto-detects the header row (single-product = row 1, multi-product = row 2 with
    an ingredient legend above it) by locating the row whose first cell is "DIN".
    Returns a DataFrame with the ``_SHEET1_KEYS`` columns that are present plus any
    IQVIA metric columns found, one row per data row.
    """
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if _SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Workbook has no {_SHEET_NAME!r} sheet (found: {wb.sheetnames})"
        )
    ws = wb[_SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()

    # Locate the header row: the first row whose first non-empty cell is "DIN".
    header_idx = None
    for i, r in enumerate(rows[:3]):
        if r and str(r[0]).strip().upper() == "DIN":
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(
            "Could not locate the header row (no leading 'DIN' cell in first 3 rows)"
        )

    raw_headers = [str(h).strip() if h is not None else "" for h in rows[header_idx]]
    header_map = _build_header_map()

    # Resolve each column to either an internal key, a metric column, or None.
    col_keys: list[Optional[str]] = []
    metric_cols: list[str] = []
    for h in raw_headers:
        if h in header_map:
            col_keys.append(header_map[h])
        elif _METRIC_COL_RE.match(h):
            col_keys.append(h)
            metric_cols.append(h)
        else:
            col_keys.append(None)

    records: list[dict] = []
    for r in rows[header_idx + 1:]:
        if r is None:
            continue
        rec: dict = {}
        for cell, key in zip(r, col_keys):
            if key is None:
                continue
            rec[key] = cell
        if not str(rec.get("din", "") or "").strip():
            continue
        records.append(rec)

    df = pd.DataFrame(records)
    # Coerce metric cells to nullable ints (blank/None stays NaN, never 0).
    for mc in metric_cols:
        if mc in df.columns:
            df[mc] = pd.to_numeric(df[mc], errors="coerce")
    return df


# ── Core verification ─────────────────────────────────────────────────────────

def _metric_tuple(row: pd.Series, metric_cols: list[str]) -> Optional[tuple[int, ...]]:
    """Return the integer metric tuple for a row, or None if every value is blank."""
    vals: list[int] = []
    any_present = False
    for c in metric_cols:
        v = row.get(c)
        if v is None or pd.isna(v):
            vals.append(0)
        else:
            vals.append(int(round(float(v))))
            any_present = True
    return tuple(vals) if any_present else None


def _is_subsequence(short: str, full: str) -> bool:
    """True when every char of ``short`` appears in ``full`` in order."""
    it = iter(full)
    return all(ch in it for ch in short)


def _is_truncated_brand(a: str, b: str) -> bool:
    """True when one brand is a character-level abbreviation of the other.

    This separates a benign *truncated* IQVIA label from a wrong *sibling* match —
    the distinction that keeps the verifier's core catch intact:

      • "mylan-efavirnz/emtr/teno" IS a subsequence of the full DIN brand
        "mylan-efavirenz/emtricitabine/tenofovir…"  → benign truncation.
      • "nova-progest" is NOT a subsequence of the sibling "nova-lutein" (and vice
        versa) — they share only the "nova" manufacturer stem but name different
        products → a genuine confident-wrong sibling swap that must stay HIGH.

    Requires the leading manufacturer/brand token to match (defence against
    coincidental short-brand subsequences) AND one alphanumeric form to be an
    ordered subsequence of the other.
    """
    aw = re.findall(r"[a-z]+", a.lower())
    bw = re.findall(r"[a-z]+", b.lower())
    if not aw or not bw or aw[0] != bw[0]:
        return False
    ca = re.sub(r"[^a-z0-9]", "", a.lower())
    cb = re.sub(r"[^a-z0-9]", "", b.lower())
    if not ca or not cb:
        return False
    return _is_subsequence(ca, cb) or _is_subsequence(cb, ca)


def _decompose_aggregate(
    wb_tuple: tuple[int, ...],
    groups: list[dict],
    s1: dict,
    metric_cols: list[str],
) -> Optional[list[dict]]:
    """Find a small set of same-strength, same-company groups summing to wb_tuple.

    The matcher aggregates a generic-label alias onto its exact-brand DIN, so a
    DIN's cell can be the sum of 2–3 groups rather than any single group's tuple.
    Candidates are restricted to the DIN's strength and a plausibly-matching
    company (so the search stays tiny — a firm rarely has >3 same-strength groups),
    and only 2- and 3-group sums are tried (the aggregation never combines more).
    Returns the matching combo, or None.
    """
    cands = [
        g for g in groups
        if g["strength_set"] == s1["strength_set"]
        and _sim(s1["company_norm"], g["company_norm"]) >= MIN_COMPANY_SIM
    ]
    n = len(cands)
    width = len(metric_cols)
    for k in (2, 3):
        if n < k:
            break
        for combo in itertools.combinations(cands, k):
            if all(
                sum(g["tuple"][i] for g in combo) == wb_tuple[i] for i in range(width)
            ):
                return list(combo)
    return None


def verify(workbook_path: str, iqvia_path: str) -> VerifyReport:
    """Verify the IQVIA matches in ``workbook_path`` against the raw IQVIA file."""
    sheet1 = _read_sheet1(workbook_path)
    with open(iqvia_path, "rb") as fh:
        collapsed = collapse_iqvia(parse_iqvia(fh.read()))

    # Metric columns common to both the workbook and the collapsed IQVIA file.
    # The export title-cases headers ("Units MAT 12/2025" → "Units Mat 12/2025"),
    # so pair them case-insensitively and rename the workbook columns back to the
    # canonical IQVIA names that the collapsed frame uses.
    iq_metrics = detect_metric_columns(collapsed)
    wb_metrics = [c for c in sheet1.columns if _METRIC_COL_RE.match(str(c))]
    upper_to_wb = {str(c).upper(): c for c in wb_metrics}
    metric_cols = [iq for iq in iq_metrics if iq.upper() in upper_to_wb]
    sheet1 = sheet1.rename(
        columns={upper_to_wb[iq.upper()]: iq for iq in metric_cols}
    )

    if sheet1.empty or collapsed.empty or not metric_cols:
        return VerifyReport([], [], metric_cols, 0, len(collapsed))

    # ── Pre-normalise the collapsed IQVIA groups ─────────────────────────────
    groups: list[dict] = []
    tuple_to_groups: dict[tuple[int, ...], list[int]] = {}
    for gi, g in collapsed.iterrows():
        gt = tuple(int(g[c]) for c in metric_cols)
        info = {
            "idx": gi,
            "product": str(g.get("Product", "") or "").strip(),
            "manufacturer": str(g.get("Manufacturer", "") or "").strip(),
            "strength": str(g.get("Strength", "") or "").strip(),
            "molecule": str(g.get("Combined Molecule", "") or "").strip(),
            "brand_norm": _norm_brand(g.get("Product", "")),
            "company_norm": _norm_company(g.get("Manufacturer", "")),
            "strength_set": _norm_strength(g.get("Strength", "")),
            "tuple": gt,
        }
        groups.append(info)
        tuple_to_groups.setdefault(gt, []).append(len(groups) - 1)

    # ── Pre-normalise the Sheet-1 rows ───────────────────────────────────────
    s1_rows: list[dict] = []
    for _, r in sheet1.iterrows():
        s1_rows.append({
            "din": str(r.get("din", "") or "").strip(),
            "brand": str(r.get("brand_name", "") or "").strip(),
            "company": str(r.get("company", "") or "").strip(),
            "strength": str(r.get("strength", "") or "").strip(),
            "brand_norm": _norm_brand(r.get("brand_name", "")),
            "company_norm": _norm_company(r.get("company", "")),
            "strength_set": _norm_strength(r.get("strength", "")),
        })

    # ── Plumbing/regression layer: re-run the live matcher and compare ───────
    matcher_input = sheet1[[k for k in _SHEET1_KEYS if k in sheet1.columns]].copy()
    canonical_enriched, _ = match_iqvia_to_sheet1(matcher_input, collapsed)
    canonical: dict[str, Optional[tuple[int, ...]]] = {}
    for _, r in canonical_enriched.iterrows():
        din = str(r.get("din", "") or "").strip()
        t = _metric_tuple(r, metric_cols)
        # A combination DIN appears in several ingredient blocks (same DIN, many
        # rows); the matcher stamps the value on the FIRST occurrence only, so the
        # later rows are blank. Keep the first non-blank tuple per DIN — otherwise a
        # trailing blank row overwrites the real value and every such combo falsely
        # trips rerun_divergence.
        if canonical.get(din) is None:
            canonical[din] = t

    # ── Per-DIN independent judgement ────────────────────────────────────────
    results: list[DinResult] = []
    for r, s1 in zip(sheet1.to_dict("records"), s1_rows):
        din = s1["din"]
        wb_tuple = _metric_tuple(pd.Series(r), metric_cols)
        if wb_tuple is None:
            continue  # DIN carries no IQVIA values — nothing to verify

        res = DinResult(
            din=din,
            brand=s1["brand"],
            company=s1["company"],
            strength=s1["strength"],
            values={c: int(round(float(r[c]))) for c in metric_cols
                    if r.get(c) is not None and not pd.isna(r.get(c))},
        )

        # Reverse-map the workbook values to the exact IQVIA group.
        cand_idx = tuple_to_groups.get(wb_tuple, [])
        assigned: Optional[dict] = None
        if len(cand_idx) == 1:
            assigned = groups[cand_idx[0]]
        elif len(cand_idx) > 1:
            # Identical sums in two groups — disambiguate by best brand match.
            assigned = max(
                (groups[i] for i in cand_idx),
                key=lambda g: _sim(s1["brand_norm"], g["brand_norm"]),
            )
            res.evidence.append(
                f"{len(cand_idx)} IQVIA groups share these exact values; "
                "resolved by best brand match"
            )

        if assigned is None:
            # The value may be an AGGREGATE: the matcher sums a generic-label alias
            # (IQVIA "METFORMIN") onto its same-company exact-brand DIN, so the cell
            # equals the sum of several same-strength, same-company groups and
            # matches no single group's tuple. Try to reconstruct that sum before
            # declaring it an orphan; anchor brand judgement on the best-brand
            # constituent (the exact-brand group, sim 100 → correctly clean).
            combo = _decompose_aggregate(wb_tuple, groups, s1, metric_cols)
            if combo:
                assigned = max(combo, key=lambda g: _sim(s1["brand_norm"], g["brand_norm"]))
                others = ", ".join(
                    f"{g['product']}" for g in combo if g["idx"] != assigned["idx"]
                )
                res.evidence.append(
                    f"value is an aggregate of {len(combo)} same-company groups "
                    f"(anchor '{assigned['product']}'"
                    + (f" + {others}" if others else "") + ")"
                )

        if assigned is None:
            # No IQVIA group (or aggregate) produces these numbers — stale workbook.
            res.flags.append("value_orphan")
            res.evidence.append("workbook values match no current IQVIA group")
            res.suspicion = _FLAG_WEIGHTS["value_orphan"]
            results.append(res)
            continue

        res.assigned_product = assigned["product"]
        res.assigned_manufacturer = assigned["manufacturer"]
        res.assigned_strength = assigned["strength"]
        b = _sim(s1["brand_norm"], assigned["brand_norm"])
        c = _sim(s1["company_norm"], assigned["company_norm"])
        res.brand_sim = round(b, 1)
        res.company_sim = round(c, 1)

        # Signal 1: company-carried (the confident-wrong fingerprint).
        if b < BRAND_FLOOR:
            res.flags.append("company_carried")
            res.evidence.append(
                f"brand_sim={b:.0f} < {BRAND_FLOOR:.0f} but company_sim={c:.0f}: "
                f"match leaned on the manufacturer, not the brand"
            )
            res.suspicion += _FLAG_WEIGHTS["company_carried"] * (BRAND_FLOOR - b)

        # Signal 2: a different Sheet-1 DIN matches this GROUP's brand better.
        # The rival DIN must plausibly belong to the GROUP's company — otherwise a
        # same-brand DIN from an unrelated firm (every generic "AMLODIPINE" DIN
        # looks alike) fires spuriously. Company is the real discriminator, so a
        # rival that cannot own this group's sales is not evidence.
        rival_din = None
        rival_din_b = b
        for other in s1_rows:
            if other["din"] == din:
                continue
            if other["strength_set"] != assigned["strength_set"]:
                continue
            if _sim(other["company_norm"], assigned["company_norm"]) < MIN_COMPANY_SIM:
                continue
            ob = _sim(other["brand_norm"], assigned["brand_norm"])
            if ob > rival_din_b:
                rival_din_b, rival_din = ob, other
        rival_din_exact = rival_din is not None and rival_din_b >= EXACT_BRAND_SIM and b < EXACT_BRAND_SIM
        if rival_din is not None and (
            (rival_din_b >= RIVAL_MIN_BRAND and (rival_din_b - b) >= RIVAL_MARGIN)
            or rival_din_exact
        ):
            res.flags.append("better_din_for_group")
            exact_note = " (exact-brand counterpart)" if rival_din_exact else ""
            res.evidence.append(
                f"DIN {rival_din['din']} ({rival_din['brand']}) matches "
                f"'{assigned['product']}' brand better: {rival_din_b:.0f} vs {b:.0f}{exact_note}"
            )
            res.suspicion += _FLAG_WEIGHTS["better_din_for_group"] * (rival_din_b - b)

        # Signal 3: a different IQVIA group matches this DIN's brand better.
        # The rival group must plausibly be this DIN's company — a same-brand group
        # from another firm (generic "AMLODIPINE" sold by everyone) is not a real
        # alternative owner of this DIN's sales.
        rival_grp = None
        rival_grp_b = b
        for g in groups:
            if g["idx"] == assigned["idx"]:
                continue
            if g["strength_set"] != s1["strength_set"]:
                continue
            if _sim(s1["company_norm"], g["company_norm"]) < MIN_COMPANY_SIM:
                continue
            gb = _sim(s1["brand_norm"], g["brand_norm"])
            if gb > rival_grp_b:
                rival_grp_b, rival_grp = gb, g
        rival_grp_exact = rival_grp is not None and rival_grp_b >= EXACT_BRAND_SIM and b < EXACT_BRAND_SIM
        if rival_grp is not None and (
            (rival_grp_b >= RIVAL_MIN_BRAND and (rival_grp_b - b) >= RIVAL_MARGIN)
            or rival_grp_exact
        ):
            res.flags.append("better_group_for_din")
            exact_note = " (exact-brand counterpart)" if rival_grp_exact else ""
            res.evidence.append(
                f"IQVIA group '{rival_grp['product']}' ({rival_grp['manufacturer']}) "
                f"matches this DIN's brand better: {rival_grp_b:.0f} vs {b:.0f}{exact_note}"
            )
            res.suspicion += _FLAG_WEIGHTS["better_group_for_din"] * (rival_grp_b - b)

        # Signal 4: the live matcher disagrees with the value in the workbook.
        if canonical.get(din) != wb_tuple:
            res.flags.append("rerun_divergence")
            exp = canonical.get(din)
            res.evidence.append(
                f"live matcher would assign {exp!r}, workbook has {wb_tuple!r}"
            )
            res.suspicion += _FLAG_WEIGHTS["rerun_divergence"]

        # Downrank the benign company-carried class. When company_carried is the
        # ONLY flag AND the assigned DIN is the single same-company candidate at this
        # strength, the match is FORCED — there is no other same-company DIN it could
        # belong to — so the weak brand leg is not evidence of a wrong match. The
        # brand is weak for one of two harmless reasons, the two false-positive
        # shapes the audit found:
        #   • a generic / USP-style IQVIA label ("AMLODIPINE BESYLATE USP")  → branded DIN, or
        #   • a truncated long combo brand ("MYLAN-EFAVIRNZ/EMTR/TENO")      → full DIN brand,
        # both of which still share the DIN's manufacturer/brand stem. Keep these
        # auditable in the low-priority tier instead of as headline noise. Genuine
        # errors always carry a second flag (better_din/better_group/value_orphan/
        # rerun_divergence) and so never reach this branch.
        if set(res.flags) == {"company_carried"}:
            same_company_candidates = sum(
                1 for o in s1_rows
                if o["strength_set"] == assigned["strength_set"]
                and _sim(o["company_norm"], assigned["company_norm"]) >= MIN_COMPANY_SIM
            )
            generic = _is_generic_brand(
                assigned["brand_norm"], assigned["molecule"], _GENERIC_FILLER_WORDS
            )
            truncated = _is_truncated_brand(assigned["brand_norm"], s1["brand_norm"])
            if same_company_candidates == 1 and (generic or truncated):
                res.tier = "low"
                why = "generic/USP-style label" if generic else "truncated brand label"
                res.evidence.append(
                    f"downranked: {why} with a single same-company candidate DIN "
                    "— forced match, weak brand leg expected"
                )

        results.append(res)

    flagged = [d for d in results if d.suspicion > 0]
    suspicious = sorted(
        (d for d in flagged if d.tier == "high"),
        key=lambda d: d.suspicion,
        reverse=True,
    )
    low_priority = sorted(
        (d for d in flagged if d.tier == "low"),
        key=lambda d: d.suspicion,
        reverse=True,
    )
    return VerifyReport(
        low_priority=low_priority,
        dins=results,
        suspicious=suspicious,
        metric_cols=metric_cols,
        n_enriched=len(results),
        iqvia_groups=len(collapsed),
    )


# ── Reporting / CLI ────────────────────────────────────────────────────────────

def format_report(report: VerifyReport) -> str:
    lines: list[str] = []
    lines.append("IQVIA match verification")
    lines.append("=" * 60)
    lines.append(f"DINs carrying IQVIA values : {report.n_enriched}")
    lines.append(f"Collapsed IQVIA groups     : {report.iqvia_groups}")
    lines.append(f"Suspicious matches         : {len(report.suspicious)}")
    lines.append(f"Low-priority (review) tier : {len(report.low_priority)}")
    lines.append("")

    def _emit(d: DinResult, i: int) -> None:
        lines.append(
            f"{i}. DIN {d.din}  suspicion={d.suspicion:.0f}  flags={','.join(d.flags)}"
        )
        lines.append(f"     Sheet-1 : {d.brand}  /  {d.company}  /  {d.strength}")
        if d.assigned_product is not None:
            lines.append(
                f"     IQVIA   : {d.assigned_product}  /  {d.assigned_manufacturer}"
                f"  /  {d.assigned_strength}   (brand_sim={d.brand_sim}, company_sim={d.company_sim})"
            )
        lines.append(f"     values  : {d.values}")
        for ev in d.evidence:
            lines.append(f"     -> {ev}")
        lines.append("")

    if not report.suspicious:
        lines.append("OK - no high-priority suspicious matches found.")
    else:
        lines.append("Ranked suspicious DINs (most suspicious first):")
        lines.append("-" * 60)
        for i, d in enumerate(report.suspicious, 1):
            _emit(d, i)

    if report.low_priority:
        lines.append("Low-priority review tier (generic/USP-style, forced single-DIN matches):")
        lines.append("-" * 60)
        for i, d in enumerate(report.low_priority, 1):
            _emit(d, i)
    return "\n".join(lines)


def _report_to_dict(report: VerifyReport) -> dict:
    def _d(x: DinResult) -> dict:
        return {
            "din": x.din, "brand": x.brand, "company": x.company,
            "strength": x.strength, "values": x.values,
            "assigned_product": x.assigned_product,
            "assigned_manufacturer": x.assigned_manufacturer,
            "assigned_strength": x.assigned_strength,
            "brand_sim": x.brand_sim, "company_sim": x.company_sim,
            "flags": x.flags, "evidence": x.evidence, "suspicion": x.suspicion,
            "tier": x.tier,
        }
    return {
        "ok": report.ok,
        "n_enriched": report.n_enriched,
        "iqvia_groups": report.iqvia_groups,
        "metric_cols": report.metric_cols,
        "suspicious": [_d(d) for d in report.suspicious],
        "low_priority": [_d(d) for d in report.low_priority],
        "dins": [_d(d) for d in report.dins],
    }


def main(argv: Optional[list[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description="Read-only post-export verifier for IQVIA→DIN matches.",
    )
    ap.add_argument("--workbook", required=True, help="finished export .xlsx")
    ap.add_argument("--iqvia", required=True, help="raw IQVIA .xlsx")
    ap.add_argument("--json", action="store_true", help="emit JSON instead of text")
    args = ap.parse_args(argv)

    report = verify(args.workbook, args.iqvia)

    if args.json:
        import json
        print(json.dumps(_report_to_dict(report), indent=2, default=str))
    else:
        print(format_report(report))

    return 0 if report.ok else 1


if __name__ == "__main__":
    sys.exit(main())
